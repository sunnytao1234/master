# -*- coding: utf-8 -*-
# @time     :   4/1/2021 3:36 PM
# @author   :   balance
# @file     :   storage.py
import os

import openpyxl
from openpyxl.styles import PatternFill

from Common.common_function import OSTool
from Common.log import log


class Capture:

    def __init__(self, status):
        """capture test for wes"""
        self.__test_data = None
        self.status = status
        self.save_path = OSTool.get_current_dir('result.xlsx')
        self.template_path = OSTool.get_current_dir('Test_Data', 'td_thinupdate', 'ThinUpdate Capture and Deploy.xlsx')
        self.data_start_col = 8
        self.data_start_row = 2
        """
        {name: 'storage', loop: 3, os: wes, platform: mt31, tester: balance@hp.com, 
        os_info: windows10-xx-xxx, platform_info: hp mt31, ml_info: xxxxx, memory_info: xxxx8G,
        disk_info: sumsun xxxxx32G, disk_pn: 12345, bios_info: xxxx, cpu_info: xxxx@2.0GHz,
        gpu_info: xxxx, main_board_info: 321a}
        """

    def check_write_filter_status(self):
        """ install passmark in uut
        1. file store in Test_utility
        2. install with quiet mode
        """
        pass

    def open_thinupdate_click_capture(self):
        """
        launch diskmark
        :return:
        """
         #to do
        pass

    def set_config(self, value: dict, **kwargs):
        self.__test_data = value
        self.save_path = OSTool.get_current_dir('Test_Report',
                                                'Thin_capture_{}_{}.xlsx'.format(value.get('platform'),
                                                                            value.get('os_info')))
    def recode_capture_start_time(self):
        """
        control to do

        :return:
        """
    def check_capture_status(self):
        """
        get diskmark test status
        :return: status of diskmark
        """

    def recode_capture_end_time(self):
        """control to do"""

        pass


    def collect_capture_result(self):
        """
        get test result and record to specific excel file
        :return: list
        return: [seq32_read, 4k32_read, seq_read, 4k_read, seq32_write, 4k32_write, seq_write, 4k_write]
        """
        pass

    def collect_deploy_result(self):
        """
        get test result and record to specific excel file
        :return: list
        return: [seq32_read, 4k32_read, seq_read, 4k_read, seq32_write, 4k32_write, seq_write, 4k_write]
        """
        pass


    def export_result_to_excle(self):
        """export collect data to excle about capture and deploy result"""
        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)
        sheet = data['Sheet1']
        max_row = sheet.max_row
        # to do
        pass

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []

        for item in result:
            if item.get('value') == 0:
                abs_value = abs(item.get('value') - average)
            else:
                abs_value = abs((item.get('value') - average) / item.get('value'))

            if abs_value > deviation:
                abnormal_positions.append(item.get('position'))

                log.debug('[Capture][get_abnormal_data]find abnormal data cell: {}'.format(item.get('position')))

        return abnormal_positions


    def analyze_report(self):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 10, row 1
        :return:
        """
        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)
        sheet = data['Storage']
        max_row = sheet.max_row
        log.debug(
            '[capture wes][analyze_report]current used rows {}, start row {}'.format(max_row, self.data_start_row))
        for col in range(self.data_start_col, self.data_start_col + 8):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
            for row in range(self.data_start_row, max_row + 1):
                cell_value = sheet.cell(row, col).value
                cell_value = float(cell_value) if cell_value is not None else 0
                sum_data += cell_value
                temp_data = {'position': (row, col), 'value': cell_value}
                result.append(temp_data)
            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row + 1 - self.data_start_row)
            sheet.cell(max_row + 1, col).value = average_data
            abnormals = self.__get_abnormal_data(result, average_data)
            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill
        data.save(self.save_path)
        data.close()

    def test_thinupdate(self):
        """
        run using specific config
        :return:
        """
        for loop in range(self.__test_data.get('loop')):
            #to do
            return True




def start(value: dict):
    """
    main function for reboot test
    1. support cross platform
    2. support all reboot test items
    3. collect result and analyze test result
       3.1 get result
       3.2 get average result
       3.3 mark abnormal data as red in excel
    4. write to excel as expected format
    5. all exception should be covered
    :param _os: thinpro7x wes10_rsx
    :param loop: test loop
    :param kwargs:
    :return: None
    """
    type = value.get('')
    # todo
    print('capture')
    return None
