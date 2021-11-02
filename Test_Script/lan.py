# -*- coding: utf-8 -*-
# @time     :   4/1/2021 3:39 PM
# @author   :   balance
# @file     :   lan.py
import os
import re
import subprocess
import time

import openpyxl
from openpyxl.styles import PatternFill

from Common.common_function import OSTool
from Common.exception import IncorrectOS
from Common.log import log
from Common.common import NetWorkDisableEnable, get_global_gateway, get_server_ip, get_lan_version


class LanTestWin:

    def __init__(self, status):
        self.__test_data = None
        self.status = status

        self.save_path = OSTool.get_current_dir('result.xlsx')
        self.template_path = OSTool.get_current_dir('Test_Data', 'excle_template.xlsx')

        self.data_start_col = 7
        self.data_start_row = 2

    def enable_wlan(self):
        NetWorkDisableEnable().enable_wlan()
        return True

    def disable_wlan(self):
        NetWorkDisableEnable().disable_wlan()
        return True

    def is_connected(gateway=get_global_gateway()) -> bool:
        """
        check whether lan is connected to network
        :return: if connnected return True, else return False
        """
        exit_code = os.system('ping {} -n 1'.format(gateway))
        log.info("[lan-wes]exit_code:{}".format(exit_code))

        if exit_code:
            log.info('[lan-wes] connect network fail')

            return False
        else:
            log.info('[lan-wes] connect network success')

            return True

    def set_config(self, value: dict, **kwargs) -> None:
        """
        set wlan test config from analyzed data
        !! must call this method before test file read and write
        :param value: lan test config,
        eg. {'name': 'wlan', 'router': 'xxx', 'platform': 'T630 DVT', 'loop': 3, 'os': 'wes10 rs5',,,,}
        :return: None
        """
        self.__test_data = value
        self.save_path = OSTool.get_current_dir('Test_Report',
                                                'Lan_{}_{}.xlsx'.format(value.get('platform'),
                                                                        value.get('os_info')))

    def test_write(self):
        """
        test file transfer using jperf from local storage to server
        :return: average speed: Mb/s
        """
        version = get_lan_version()
        uut_iperf_path = OSTool.get_current_dir('Test_Utility', '{}'.format(version))
        cmd = "{tool} -c {server_ip} -t 120 ".format(tool=uut_iperf_path, server_ip=get_server_ip())
        uut_output = subprocess.Popen(cmd, shell=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        log.info("[lan_wes]uut_output value is:{}".format(uut_output))
        lines = uut_output.stdout.readlines()
        log.info("[lan_wes] lines value:{}".format(lines))
        test_write_value = re.findall(r"\d+\.?\d*", str(lines[-3]))
        log.info('[lan-wes][test_write] test_write value:{}'.format(test_write_value[-1]))

        return test_write_value[-1]

    def collect_result(self) -> None:
        """
        get test result and record to specific excel file
        :return:
        """
        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)

        test_write = self.test_write()
        lan_list = ['', '', self.__test_data.get('platform'),
                    self.__test_data.get('bios_info'), self.__test_data.get('ml_info'), self.__test_data.get('os_info'),
                    '{test_write}'.format(test_write=test_write)]

        sheet = data['LAN']
        max_row = sheet.max_row
        log.info("mac row value is :{}".format(max_row))

        for i in range(len(lan_list)):
            sheet.cell(max_row + 1, i + 1).value = lan_list[i]

        data.save(self.save_path)
        data.close()

    def test_lan(self):
        """
          run using specific config
          :return:
        """
        log.info('[lan][lantestwin]begin to test')
        for loop in range(self.__test_data.get('loop')):
            self.collect_result()
            log.info(f'[lan wes][test lan]Test cycle {loop + 1} Finished')

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []
        for item in result:
            if abs(item.get('value') - average) / item.get('value') > deviation:
                abnormal_positions.append(item.get('position'))
                log.debug('[storage][get abnormal data]Find abnormal data cell: {}'.format(item.get('position')))
        return abnormal_positions

    def analyze_report(self):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 7, row 2
        :return:
        """
        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)
        sheet = data['LAN']
        max_row = sheet.max_row
        for col in range(self.data_start_col, self.data_start_col + 1):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
            for row in range(self.data_start_row, max_row + 1):
                cell_value = sheet.cell(row, col).value
                cell_value = float(cell_value) if cell_value is not None else 0
                sum_data += cell_value
                temp_data = {'position': (row, col), 'value': cell_value}
                result.append(temp_data)
            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row + 1 - self.data_start_row)
            sheet.cell(max_row + 1, col).value = average_data
            abnormals = self.__get_abnormal_data(result, average_data)
            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill
        data.save(self.save_path)
        data.close()


class LanTestLinux:
    def __init__(self, status):
        self.__test_data = None
        self.status: dict = status
        self.save_path = OSTool.get_current_dir('result.xlsx')
        self.template_path = OSTool.get_current_dir('Test_Data', 'excle_template.xlsx')
        self.data_start_col = 7
        self.data_start_row = 2

    def enable_wlan(self) -> bool:
        subprocess.Popen('mclient --quiet set root/Network/Wireless/EnableWireless 1 && mclient commit', shell=True,
                         stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        value = subprocess.Popen("mclient --quiet get root/Network/Wireless/EnableWireless", shell=True,
                                 stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        s = value.stdout.read().decode()
        if str(s[0]) == '1':
            return True
        else:
            return False

    def disable_wlan(self) -> bool:
        subprocess.Popen('mclient --quiet set root/Network/Wireless/EnableWireless 0 && mclient commit', shell=True,
                         stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        value = subprocess.Popen("mclient --quiet get root/Network/Wireless/EnableWireless", shell=True,
                                 stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        s = value.stdout.read().decode(encoding='utf-8')
        if str(s) == '':
            return True
        else:
            return False

    def set_config(self, value: dict, **kwargs) -> None:
        """
        set lan test config from analyzed data
        !! must call this method before test file read and write
        :param value: lan test config,
        eg. {'name': 'lan',, 'platform': 'T630', 'loop': 3, 'os': 'linux',,,,}
        :return: None
        """
        self.__test_data = value
        self.save_path = OSTool.get_current_dir('Test_Report',
                                                'lan_linux_{}_{}.xlsx'.format(value.get('platform'),
                                                                              value.get('os_info')))

    @property
    def is_connected(self) -> bool:
        """
        check whether lan is connected to network
        :return: if connnected return True, else return False
        """
        exit_code = os.system('ping {} -n 1'.format(self.__test_data.get('network').get('gateway')))
        log.info("[lan-wes]exit_code:{}".format(exit_code))
        if exit_code:
            log.info('[lan-wes] connect network fail')
            return False
        else:
            log.info('[lan-wes] connect network success')
            return True

    def install(self):
        """
        test file transfer using jperf from local storage to server
        :return: average speed: Mb/s
        """
        lib_version, jper_version = get_lan_version()
        tool_path = os.path.abspath(OSTool.get_current_dir("Test_Utility"))
        os.system('chmod 777 {}'.format(tool_path))
        log.info("tool_path value:{}".format(tool_path))
        os.system('fsunlock')
        install_liperf = subprocess.getoutput('sudo dpkg -i {a}/{b}'.format(a=tool_path, b=lib_version))
        log.info("install lib result is:{}".format(install_liperf))
        os.system('fsunlock')
        install_jper = subprocess.getoutput('sudo dpkg -i {a}/{b}'.format(a=tool_path, b=jper_version))
        log.info("install jperf result is:{}".format(install_jper))
        check_iperf3 = subprocess.getoutput('sudo dpkg -l | grep iperf3')
        log.info("[lan-linux][install]check_iperf3 value:{}".format(check_iperf3))
        if check_iperf3 == '':
            log.error("[lan-linux][install]install iperf3 fail")
            return False
        else:
            log.info("[lan-linux][install]install iperf3 successful")
            return True

    def test_write(self):
        log.info("[lan-linux][test_write]start to test lan")
        iperf3_run_path = subprocess.getoutput('which iperf3').strip().split('/iperf3')[0]
        log.info("[lan-linux][test_write]iperf3_run_path value:{}".format(iperf3_run_path))
        cmd = " sudo /usr/bin/iperf3 -c {server_ip} -t 120 -w 56.0m".format(server_ip=get_server_ip()).replace('\n', '')
        uut_linux_output = subprocess.getoutput(cmd)
        log.info("[lan_linux][test_write]uut_linux_output value:{}".format(uut_linux_output))
        test_write_value = re.findall(r"\d+\.?\d*", str(uut_linux_output).strip().
                                      split('- - - - - - - - - - - - - - - - - - - - - - - - -')[1])
        log.info('[lan-wes][test_write] test_write value:{}'.format(test_write_value[-1]))
        return test_write_value[-1]

    def collect_result(self) -> None:
        """
        get test result and record to specific excel file
        :return:
        """
        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)
        test_write = self.test_write()
        version = subprocess.getoutput('cat /etc/imageid')
        lan_list = ['', '', self.__test_data.get('platform_info'),
                    self.__test_data.get('bios_info'), self.__test_data.get('ml_info'), self.__test_data.get('os_info'),
                    '{test_write}'.format(test_write=test_write)]
        sheet = data['LAN']
        max_row = sheet.max_row
        log.info("mac row value is :{}".format(max_row))
        for i in range(len(lan_list)):
            sheet.cell(max_row + 1, i + 1).value = lan_list[i]
        data.save(self.save_path)
        data.close()

    def test_lan(self):
        """
          run using specific config
          :return:
        """
        log.info('[lan linux][test lan]begin to test')
        for loop in range(self.__test_data.get('loop')):
            self.collect_result()
            log.info(f'[lan linux][test lan]Test cycle {loop + 1} Finished')

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []
        for item in result:
            if item.get('value') == 0:
                abs_value = abs(item.get('value') - average)
            else:
                abs_value = abs((item.get('value') - average) / item.get('value'))
            if abs_value > deviation:
                abnormal_positions.append(item.get('position'))
                log.debug('[storage][get abnormal data]Find abnormal data cell: {}'.format(item.get('position')))
        return abnormal_positions

    def analyze_report(self):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 7, row 2
        :return:
        """
        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)
        sheet = data['LAN']
        max_row = sheet.max_row
        for col in range(self.data_start_col, self.data_start_col + 1):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
            for row in range(self.data_start_row, max_row + 1):
                cell_value = sheet.cell(row, col).value
                cell_value = float(cell_value) if cell_value is not None else 0
                sum_data += cell_value
                temp_data = {'position': (row, col), 'value': cell_value}
                result.append(temp_data)
            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row + 1 - self.data_start_row)
            sheet.cell(max_row + 1, col).value = average_data
            abnormals = self.__get_abnormal_data(result, average_data)
            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill
        data.save(self.save_path)
        data.close()


def start(value: dict, **kwargs):
    """
    main function for lan test
    1. support cross platform
    2. support all lan test items
    3. collect result and analyze test result
       3.1 get result
       3.2 get average result
       3.3 mark abnormal data as red in excel
    4. write to excel as expected format
    5. all exception should be covered
    :param _os: thinpro7x wes10_rsx
    :param loop: test loc
    :param kwargs:
    :return: None
    """
    log.debug('[lan][start]begin to test')
    status = kwargs.get('status')
    try:
        if value.get('os') == 'wes':
            lan = LanTestWin(status)
        elif value.get('os') == 'linux':
            lan = LanTestLinux(status)
            lan.install()
        else:
            log.error('[lan][start]Incorrect OS given')
            raise IncorrectOS('Incorrect OS given')
        lan.disable_wlan()
        lan.set_config(value)
        lan.test_lan()
        lan.analyze_report()
        time.sleep(10)
        lan.enable_wlan()

    except:
        import traceback
        traceback.print_exc()
        log.error('[lan][start]Get exception during test lan')
    return None
