import os
import platform
import re
import subprocess
import time
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List

import openpyxl

if 'windows' in platform.platform().lower():
    pass

from openpyxl.styles import PatternFill

from Common.log import log
from Common.tc_info import WESInfo
from Common.common import OSTool, get_cpu_core, get_ml, get_element, get_current_dir


class Passmark:
    # todo:ask Sunny why to code like this
    passmark_exe_path = r'C:\Program Files\PerformanceTest\PerformanceTest64.exe'
    unstall_passmark = r'C:\Program Files\PerformanceTest\unins000.exe'
    check_path = "C:\Program Files\PerformanceTest\PerformanceTest64.exe"

    def __init__(self):
        """performance test tool for wes"""
        self.__test_data = None

        self.data_start_row = 2
        self.data_start_col = 12
        self.data_end_row = 5

        self.save_path = OSTool.get_current_dir('result.xlsx')
        self.template_path = OSTool.get_current_dir('Test_Data', 'excle_template.xlsx')

    def install(self):
        """ install passmark in uut
        1.TXT. file store in Test_utility
        2. install with quiet mode
        """
        if os.path.exists(self.passmark_exe_path):
            log.info('[passmark][install] this computer already install passmark')

            return True

        passmark_packet_path = OSTool.get_current_dir("Test_Utility", "petst.exe")
        subprocess.Popen('{} /silent'.format(passmark_packet_path))
        time.sleep(120)

        if os.path.exists(self.check_path):
            log.info('[Passmark][install]install passmark successful')

            return True

        else:
            log.error('[Passmark][install]install passmark fail')

            return False

    def install_license(self, check_cycle: int = 3, wait_time: int = 10):
        """
        license store in Test utility(txt file)
        :param license_value:
        :return:
        """
        subprocess.Popen(self.passmark_exe_path)
        time.sleep(wait_time)

        for i in range(check_cycle):
            license_path = OSTool.get_current_dir("Test_Utility", "Key.txt")
            f = open(license_path)

            license_all_value = f.read()
            log.info('[Passmark][install_license] licese_all_value:{}'.format(license_all_value))
            license_value = \
                license_all_value.strip().split('-----START_OF_KEY-----')[1].strip().split('------END_OF_KEY------')[
                    0]
            log.info('[Passmark][install_license] licese_value:{}'.format(license_value))
            f.close()

            main_window = get_element("MAIN_WINDOW")
            # todo: fix performance issue
            if main_window.Exists() and not get_element("MAIN_SECOND_WINDOW").Exists():
                log.info('[Passmark][install_license]this license have already install ')
                self.stop()

                return True

            else:
                main_second_window = get_element("MAIN_SECOND_WINDOW")
                continue_btn = get_element("CONTINUE")
                thanks_window = get_element("CONTINUE")
                ok_btn = get_element("OK")

                if not main_second_window.Exists():
                    log.error("[Passmark][install_license]not found main_second_windows")
                    time.sleep(wait_time)

                    continue

                else:
                    main_second_window.Click()
                    main_second_window.GetChildren()[2].SetValue(license_value)
                    time.sleep(wait_time)

                    if not continue_btn.Exists():
                        log.error("[Passmark][install_license]not appear continue btn")
                        self.stop()

                        return False

                    continue_btn.Click()
                    if not thanks_window.Exists():
                        self.stop()
                        log.error("[Passmark][install_license]not appear thanks window")

                        return False

                    ok_btn.Click()
                    self.stop()

                    return True

        log.debug("[Passmark][install_license]not found main second window")
        self.stop()

        return True

    def launch(self, check_cycle: int = 30, wait_time: int = 10):
        # todo: change implement methods, so this function is useless
        """
        launch passmark tool and run using specific config
        there are 2 method to run tool: run using UI and quiet mode(run with scripts file)
        :return:
        """

        # subprocess.Popen(self.passmark_exe_path)
        # time.sleep(wait_time)
        #
        # for i in range(check_cycle):
        #     # main_windows = get_element("MAIN_WINDOW")
        #     main_windows = get_ui_item(uiauto.WindowControl, name="PerformanceTest 10.0", search_depth=10)
        #     run_all_test = get_element("RUN_ALL_BENCHMARK_TESTS")
        #     yes_btn = get_element("YES")
        #
        #     if not main_windows.Exists():
        #         log.error("[Passmark][launch]not found passmark main windows")
        #         time.sleep(wait_time)
        #
        #         continue
        #
        #     else:
        #         time.sleep(30)  # todo:change time
        #         main_windows.Maximize()
        #         log.info('[Passmark][launch]begin to test performance wes')
        #
        #         # todo: fix issue
        #         # main_windows.GetLastChildControl().GetChildren()[3].Click()
        #         # time.sleep(wait_time)
        #         # log.info('[Passmark][launch]click btn value:{}'.format(
        #         #     main_windows.GetLastChildControl().GetChildren()[3].Name))
        #         # tests = get_element("TESTS")
        #         # tests.GetNextSiblingControl().Expand()
        #         # run_all_test_btn = get_element("RUN_TEST")
        #         # run_all_test_btn.Click()
        #
        #         tests_item: uiauto.MenuItemControl = get_ui_item(main_windows.MenuItemControl, name="Tests",
        #                                                          search_depth=10)
        #         tests_item.Expand()
        #         tests_menu: uiauto.MenuControl = get_ui_item(main_windows.MenuControl, name="Tests",
        #                                                      search_depth=10)
        #         run_all_test_menu_item: uiauto.MenuItemControl = get_ui_item(tests_menu.MenuItemControl,
        #                                                                      name="Run All Tests",
        #                                                                      search_depth=10)
        #         run_all_test_menu_item.Click()
        #
        #         if not run_all_test.Exists():
        #             log.error("[Passmark][launch]not appear test window")
        #
        #             continue
        #
        #         yes_btn.Click()
        #         time.sleep(720)
        #
        #         return True
        #
        # log.error("[Passmark][launch]found main windows time out")
        # self.stop()
        pass
        return False

    def set_config(self, value: dict, **kwargs):
        self.__test_data = value
        self.report_path = "C:/passmark_results.txt"
        self.save_path = OSTool.get_current_dir('Test_Report',
                                                'Passmark_{}_{}.xlsx'.format(value.get('platform_info'),
                                                                             value.get('os_info')))

    def check_status(self, check_cycle: int = 40, wait_time: int = 10) -> str:
        """
        get passmark test status
        :return:
        """
        # for i in range(check_cycle):
        #     main_window = get_element("MAIN_WINDOW")
        #     update_window = get_element("UPDATE")
        #     exit = get_element("EXIT")
        #     ok_btn = get_element("OK_BUTTON")
        #     mermoy_warning = get_element('WARNING')
        #     yes_btn = get_element('YES')
        #
        #     if mermoy_warning.Exists():
        #         value = mermoy_warning.GetChildren()[-2].Name
        #         log.warning('[Passmark][check_status]warning windows value is:{}'.format(value))
        #         log.warning('[Passmark][check_status]run memory have appear some warning',
        #                     OSTool.get_current_dir('Test_Report', 'img',
        #                                            'mermory_warning_to_passmark.jpg{}'))
        #         yes_btn.Click()
        #         time.sleep(wait_time * 36)
        #     if not ok_btn.Exists():
        #         log.info('[passmark][check_status]not found ok_btn')
        #         time.sleep(wait_time)
        #
        #         continue
        #
        #     else:
        #         main_window.GetFirstChildControl().Click()
        #         ok_btn.Click()
        #         time.sleep(wait_time)
        #
        #         if not update_window.Exists():
        #             log.error("[Passmark][check status] not found update_windows btn")
        #             self.stop()
        #
        #             return False
        #
        #         exit.Click()
        #         status = 'Completed'
        #
        #         return status
        #
        # log.info('[passmark][check_status]found ok_btn time out')
        # status = 'Running_time_out'

        status = "Completed"

        return status

    def stop(self):
        os.system('taskkill /F /IM PerformanceTest64.exe')

    def export_txt(self, wait_time: int = 600):
        """
         export test result
         :return:
         """
        # todo: change implement methods by passmark auto-launch script
        # if os.path.exists(self.report_path):
        #     os.remove(self.report_path)
        #
        # for i in range(check_cycle):
        #
        # main_windows = get_element("MAIN_WINDOW")
        # main_windows: uiautomation.WindowControl = get_ui_item(uiautomation.WindowControl,
        #                                                        name="PerformanceTest 10.0", search_depth=5)
        # file_menu = get_element("FILE")
        # export_btn = get_element("EXPORT")
        # edit_path = get_element("EDIT_PATH")
        # warning_btn = get_element("WARNING")
        # yes_btn = get_element("YES")
        #
        # if not main_windows.Exists():
        #     log.error("[Passmark][export_txt] not found main_windows")
        #     time.sleep(wait_time)
        #     continue
        #
        # main_windows.GetLastChildControl().GetChildren()[0].Expand()
        # file_menu: uiautomation.MenuControl = get_ui_item(main_windows.MenuControl, name="File", search_depth=10)
        # file_menu.GetChildren()[0].SetFocus()
        #
        # for i, item in enumerate(file_menu.GetChildren()):
        #     item.SetFocus()
        #     log.warning(f"[{i}]{item.Name} {item.AutomationId} {item.AccessibleCurrentChildId()}")
        #
        # file_menu.GetChildren()[6].Click()
        # save_text_menuitem: uiautomation.MenuItemControl = get_ui_item(main_windows.MenuItemControl,
        #                                                                name="Save results as text...",
        #                                                                search_depth=10)

        # if not edit_path.Exists():
        #     log.error('[Passmark][export_txt]not found edit window')
        #     return False
        #
        # edit_path.SetValue('{}'.format(self.report_path))
        # time.sleep(wait_time)
        #
        # export_btn.Click()
        # if warning_btn.Exists():
        #     yes_btn.Click()
        #
        # time.sleep(wait_time)
        # self.stop()

        # passmark_main_window: uiauto.WindowControl = get_ui_item(uiauto.WindowControl,
        #                                                          name="PerformanceTest 10.0",
        #                                                          search_depth=10)
        # if not passmark_main_window.Exists():
        #     log.error("[Passmark][export_txt] not found passmark window")
        #     time.sleep(wait_time)
        #     continue
        #
        # passmark_main_window.GetLastChildControl().GetChildren()[0].Expand()
        # file_menu: uiauto.MenuControl = get_ui_item(uiauto.MenuControl, name="File", search_depth=10)
        # save_as_web_item: uiauto.MenuItemControl = file_menu.GetChildren()[6]
        # save_as_web_item.Click()
        #
        # save_ok_btn: uiauto.ButtonControl = get_ui_item(uiauto.ButtonControl, name="OK", search_depth=10)
        # save_ok_btn.Click()
        # save_name_edit: uiauto.EditControl = get_ui_item(uiauto.EditControl, name="File name:", search_depth=10)
        # save_name_edit.SetValue('{}'.format(self.report_path))
        # save_btn: uiauto.ButtonControl = get_ui_item(uiauto.ButtonControl, name="Save", search_depth=10)
        # save_btn.Click()
        status = False

        if os.path.exists(self.report_path):
            os.remove(self.report_path)

        # fix path issue
        autorun_cmd = r'"C:\Program Files\PerformanceTest\PerformanceTest64.exe"' + " /s " + get_current_dir(
            'Test_Data', 'passmark', 'autotest.ptscript')
        log.info(f"[Passmark][export_txt]this command will be executed: {autorun_cmd}")
        subprocess.getoutput(autorun_cmd)

        for i in range(50):
            if os.path.exists(self.report_path):
                status = True
                break
            else:
                status = False
                time.sleep(i)
                continue
        else:
            status = False
            log.error(f"[Passmark][export_txt]cannot get passmark results from: {self.report_path}")

        return status

    def collect_result(self, check_cycle: int = 30, wait_time: int = 2) -> None:
        """
        get test result and record to specific excel file
        :return:
        """
        new_read_txt_keys = []
        new_read_txt_values = []

        for i in range(check_cycle):
            if os.path.exists(self.report_path):
                log.info('[Passmark][collect_result]export txt document is exists')

                break

            else:
                time.sleep(wait_time)
                log.debug("[Passmark][collect_result] not found txt document")

                continue

        with open(self.report_path, 'r') as read_txt:
            read_txt_list: List[str] = read_txt.readlines()
            log.info('[Passmark][collect_result]read_txt_list value:{}'.format(read_txt_list))

        for data in read_txt_list:
            if 'http://www.passmark.com' in data.strip() or not ':' in data.strip():
                log.debug("[Passmark][collect_result]not recorde data value:{}".format(data.strip()))
            else:
                key = data.strip().split(':')[0].strip()
                log.info('[Passmark][collect_result]key value:{}'.format(key))
                value1 = data.strip().split(':')[1].strip()
                log.info('[Passmark][collect_result]value1:{}'.format(key))

                if value1.endswith(')'):
                    value = value1.split('(')[0].strip()
                else:
                    value = value1

                new_read_txt_keys.append(key)
                new_read_txt_values.append(value)

        log.info(f'[Passmark][collect_result]new_read_txt_keys value{new_read_txt_keys}')
        log.info(f'[Passmark][collect_result]new_read_txt_values value{new_read_txt_values}')
        result_dict: Dict[str, str] = dict(zip(new_read_txt_keys, new_read_txt_values))
        log.info(f'[Passmark][collect_result] result_dict value {result_dict}')

        cpu_core = WESInfo().cpu_info[0].get('CpuCores')
        cpu_version = self.__test_data.get('cpu_info')
        replace_value = '{}C@'.format(cpu_core)
        cpu_type = cpu_version.replace('@', replace_value)
        log.info('[Passmark][collect_result]cpu_value:{}'.format(cpu_type))

        dx12_value = result_dict.get('DirectX 12 (Frames/Sec.)')
        if dx12_value is None:
            dx12_value = 'N/A'
        else:
            dx12_value = dx12_value

        passmark_list = ['', self.__test_data.get('platform_info'), '',
                         '{a}[{b}]'.format(a=self.__test_data.get('os_info'), b=get_ml()),
                         self.__test_data.get('bios_info'), self.__test_data.get('cpu_info'),
                         result_dict.get('Motherboard'), self.__test_data.get('memory_info'),
                         result_dict.get('Videocard'),
                         result_dict.get('Hard Drive'), '', result_dict.get('Integer Math (MOps./Sec.)'),
                         result_dict.get('Floating Point Math (MOps./Sec.)'),
                         result_dict.get('Prime Numbers (Million Primes/Sec.)'),
                         result_dict.get('Extended Instructions (SSE) (Mill. Matrices/Sec.)'),
                         result_dict.get('Compression (KBytes/Sec.)'),
                         result_dict.get('Encryption (MBytes/Sec.)'), result_dict.get('Physics (Frames/Sec.)'),
                         result_dict.get('Sorting (Thousand Strings/Sec.)'),
                         result_dict.get('CPU Single Threaded (MOps./Sec.)'),
                         result_dict.get('Cross-platform Mark (Composite average)'),
                         result_dict.get('Simple Vectors (Thousand Vectors/Sec.)'),
                         result_dict.get('Fonts and Text (Ops./Sec.)'),
                         result_dict.get('Windows Interface (Ops./Sec.)'),
                         result_dict.get('Image Filters (Filters/Sec)'),
                         result_dict.get('Image Rendering (Thousand Images/Sec)'),
                         result_dict.get('Direct 2D (Frames/Sec.)'),
                         result_dict.get('PDF Rendering (Ops./Sec.)'), result_dict.get('Direct 2D - SVG (Frames/Sec.)'),
                         result_dict.get('DirectX 9 (Frames/Sec.)'),
                         result_dict.get('DirectX 10 (Frames/Sec.)'), result_dict.get('DirectX 11 (Frames/Sec.)'),
                         dx12_value,
                         result_dict.get('GPU Compute (Ops./Sec.)'),
                         result_dict.get('Database Operations (KOps./Sec.)'),
                         result_dict.get('Memory Read Cached (MBytes/Sec.)'),
                         result_dict.get('Memory Read Uncached (MBytes/Sec.)'),
                         result_dict.get('Memory Write (MBytes/Sec.)'), result_dict.get('Available RAM (Megabytes)'),
                         result_dict.get('Memory Latency (ns (lower is better))'),
                         result_dict.get('Memory Threaded (MBytes/Sec.)'),
                         result_dict.get('Disk Sequential Read (MBytes/Sec.)'),
                         result_dict.get('Disk Sequential Write (MBytes/Sec.)'),
                         result_dict.get('IOPS 32KQD20 (MBytes/Sec.)'), result_dict.get('IOPS 4KQD1 (MBytes/Sec.)'),
                         result_dict.get('CPU Mark (Composite average)'),
                         result_dict.get('2D Graphics Mark (Composite average)'),
                         result_dict.get('Memory Mark (Composite average)'),
                         result_dict.get('Disk Mark (Composite average)'),
                         result_dict.get('3D Graphics Mark (Composite average)'),
                         result_dict.get('PassMark Rating (Composite average)')]
        read_txt.close()

        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)

        sheet = data['Passmark 10']
        max_row = sheet.max_row
        log.info("[Passmark][collect]max row value is :{}".format(max_row))

        for i in range(len(passmark_list)):
            sheet.cell(max_row + 1, i + 1).value = passmark_list[i]

        data.save(self.save_path)
        data.close()

        if os.path.exists(self.report_path):
            os.remove(self.report_path)

        time.sleep(3)
        log.info('[Passmark][collect_result] collect_result ended')

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []

        for item in result:
            if item.get('value') == 0:
                abs_value = abs(item.get('value') - average)
            else:
                abs_value = abs((item.get('value') - average) / item.get('value'))

            if abs_value > deviation:
                abnormal_positions.append(item.get('position'))

                log.debug('[Passmark][get_abnormal_data]find abnormal data cell: {}'.format(item.get('position')))

        return abnormal_positions

    def test_performance(self):
        log.info('[Passmark][test_passmark]begin to test')
        for loop in range(self.__test_data.get('loop')):
            self.launch()
            status = self.check_status()
            log.info(f'[passmark][test_performance] status value {status} ')

            if not status == 'Completed':
                log.error('f[passmark][test_performance] status value {status}')

                return False

            else:
                self.export_txt()
                self.collect_result()
                log.info(f'[Passmark][test_passmark]Test cycle {loop + 1} Finished')

    def analyze_report(self):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 10, row 1.
        :return:
        """
        global answer_num

        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)

        sheet = data['Passmark 10']
        max_row = self.data_end_row
        max_col = sheet.max_column
        log.info("[passmark][analyze_report] max_row value:{a},max_col value:{b}".format(a=max_row, b=max_col))
        log.info(
            '[passmark][analyze_report]current used rows {}, start row {}'.format(max_row, self.data_start_col))

        for col in range(self.data_start_col, max_col + 1):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]

            for row in range(self.data_start_row, max_row):
                cell_value = sheet.cell(row, col).value
                if cell_value == 'N/A':
                    recode = 'N/A'
                    cell_value = 0
                elif cell_value == 0:
                    recode = 0
                    cell_value = 0
                else:
                    recode = float(cell_value)
                    cell_value = float(cell_value)

                sum_data += cell_value
                temp_data = {'position': (row, col), 'value': cell_value}
                result.append(temp_data)

            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row - self.data_start_row)

            origin_num = Decimal(average_data)
            answer_num = origin_num.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

            if answer_num == 0 and recode == 'N/A':
                answer_num = 'N/A'
            elif answer_num == 0 and recode == 0:
                answer_num = 0
            else:
                answer_num = answer_num
            sheet.cell(max_row, col).value = answer_num

            sheet.cell(5, max_col + 1).value = 'Average_Data'.title()
            abnormals = self.__get_abnormal_data(result, average_data)

            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill

        data.save(self.save_path)
        data.close()


class Greekbech:
    def __init__(self):
        """performance test tool for linux"""
        self.__test_data = None

        self.data_start_row = 3
        self.data_start_col = 11

        self.save_path = OSTool.get_current_dir('result.xlsx')
        self.template_path = OSTool.get_current_dir('Test_Data', 'excle_template.xlsx')

    def install(self) -> bool:
        """install geekbench
        1.TXT. change geekbench5 permission 777
        2. change geekbench_x86_64 permission 777
        """
        status = False

        install_path_document = os.path.abspath(OSTool.get_current_dir("Test_Utility", "Geekbench-5.2.0-Linux"))
        log.info('[Greekbech][install] install_path_document value:{}'.format(install_path_document))

        subprocess.getoutput('fsunlock')
        os.system('chmod -R 777 {}'.format(install_path_document))

        # os.chdir(install_path_document)
        # os.system('pwd')

        install_geekbench = subprocess.getoutput(get_current_dir("Test_Utility", "Geekbench-5.2.0-Linux", 'GB.sh'))
        log.info('[Greekbech][install] install_geekbench value:{}'.format(install_geekbench))

        if 'Thank you for purchasing Geekbench' in install_geekbench:
            log.info('[Greekbech][install] install greekbench successful')
            status = True
        else:
            log.error('[Greekbech][install] install greekbench fail')
            status = False

        return status

    def install_license(self):
        """
        geekbench license already in shell command
        here will not install
        :return:
        """
        pass

    def launch(self):
        """
        launch geekbench using shell command
        geekbench5 -r gary.walker@hp.com W5Y42-LA44D-AC5ER-HJR5O-GW4HI-KYGLY-HZGYU-TOWDL-4GEB2
        need to research geekbench5 parameter
        :return:
        """
        log.info('[Greekbech][launch] start collect_result')
        geekbench_path_document = os.path.abspath(
            OSTool.get_current_dir("Test_Utility", "Geekbench-5.2.0-Linux", "geekbench5"))
        log.info('[Greekbech][launch] geekbench path value:{}'.format(geekbench_path_document))
        current_path = os.getcwd()
        log.info('[Geekbench][launch] current path value:{}'.format(current_path))
        out_put = subprocess.Popen(geekbench_path_document + ' --save result --upload', shell=True,
                                   stdout=subprocess.PIPE,
                                   stderr=subprocess.STDOUT)
        log.info('[Greekbech][launch] out_put value:{}'.format(out_put.stdout))
        out_put_value = out_put.stdout.readlines()
        log.info('[Greekbech][launch] out_put_value value:{}'.format(out_put_value))

        keys_sys, values_sys, keys_processor, values_processor, memory_keys, memory_values, sing_keys, sing_values, \
        multi_keys, multi_values, bench_keys, bench_values = [], [], [], [], [], [], [], [], [], [], [], []
        sys_start = pro_start = mer_start = multi_start = ben_start = single_start = end_number = 0
        sys_info, process_info, mer_info, single_core, multi_core, ben_info = {}, {}, {}, {}, {}, {}
        length = len(out_put_value)
        log.info('[Greekbech][collect_result] length value:{}'.format(length))

        for i in range(length):
            data = out_put_value[i].decode().strip()
            if data == 'System Information':
                sys_start = i
            elif data == 'Processor Information':
                pro_start = i
            elif data == 'Memory Information':
                mer_start = i
            elif data == 'Single-Core':
                single_start = i
            elif data == 'Multi-Core':
                multi_start = i
            elif data == 'Benchmark Summary':
                ben_start = i
            elif 'Floating Point Score' in data:
                end_number = i

        for i in range(length):
            data = out_put_value[i].decode().strip()
            re_value = re.sub('\s\s+', ':', data)

            if i > sys_start and i < pro_start - 1:
                key = re_value.split(':')[0]
                value = re_value.split(':')[1]
                keys_sys.append(key)
                values_sys.append(value)
            elif i > pro_start and i < mer_start - 1:
                key = re_value.split(':')[0]
                value = re_value.split(':')[1]
                keys_processor.append(key)
                values_processor.append(value)
            elif i > mer_start and i < single_start - 2:
                key = re_value.split(':')[0]
                value = re_value.split(':')[1]
                memory_keys.append(key)
                memory_values.append(value)
            elif i > single_start and i < multi_start - 1:
                key = re_value.split(':')[0]
                value = re_value.split(':')[1]
                sing_keys.append(key)
                sing_values.append(value)
            elif i > multi_start and i < ben_start - 1:
                key = re_value.split(':')[0]
                value = re_value.split(':')[1]
                multi_keys.append(key)
                multi_values.append(value)
            elif i > ben_start and i < end_number + 1:
                key = re_value.split(':')[0]
                value = re_value.split(':')[1]
                bench_keys.append(key)
                bench_values.append(value)

        sys_info.update(dict(zip(keys_sys, values_sys)))
        log.info('[Greekbech][collect_result]sys_info value:{}'.format(sys_info))
        process_info.update(dict(zip(keys_processor, values_processor)))
        log.info('[Greekbech][collect_result]process_info value:{}'.format(process_info))
        mer_info.update(dict(zip(memory_keys, memory_values)))
        log.info('[Greekbech][collect_result]mer_info value:{}'.format(mer_info))
        single_core.update(dict(zip(sing_keys, sing_values)))
        log.info('[Greekbech][collect_result]Single_Core value:{}'.format(single_core))
        multi_core.update(dict(zip(multi_keys, multi_values)))
        log.info('[Greekbech][collect_result]Multi_Core value:{}'.format(multi_core))
        ben_info.update(dict(zip(bench_keys, bench_values)))
        log.info('[Greekbech][collect_result]Benchmark_Summary value:{}'.format(ben_info))
        return sys_info, process_info, mer_info, single_core, multi_core, ben_info

    def set_config(self, value: dict, **kwargs):
        self.__test_data = value
        self.save_path = OSTool.get_current_dir('Test_Report',
                                                'Geekbench5_{}_{}.xlsx'.format(value.get('platform_info'),
                                                                               value.get('os_info')))

    def check_status(self):
        pass

    def stop(self):
        pass

    def collect_result(self):
        """
        get test result and record to specific excel file
        :return:
        """
        sys_info, process_info, mer_info, single_core, multi_core, ben_info = self.launch()
        get_sp = subprocess.getoutput("dpkg -l | grep hptc-sp-thinpro.*-sp")
        sp_version = re.findall('(?i) (\d+.\d+) ', get_sp, re.S)

        if not sp_version:
            sp = 0
        else:
            sp = sp_version[0]

        log.info('[Geekbench][collect_result] sp value:{}'.format(sp))
        cpu_core = get_cpu_core()
        cpu_version = self.__test_data.get('cpu_info')
        replace_value = '{}C @'.format(cpu_core)
        cpu_type = cpu_version.replace('@', replace_value)
        log.info('[Geekbench][collect_result]cpu_value:{}'.format(cpu_type))
        geekbench_list = ['', self.__test_data.get('platform_info'), '',
                          '{a}[{b}+{c}]'.format(a=self.__test_data.get('os_info'),
                                                b=subprocess.getoutput('cat /etc/imageid'),
                                                c=sp),
                          self.__test_data.get('bios_info'), self.__test_data.get('cpu_info'),
                          sys_info.get('Motherboard'), self.__test_data.get('memory_info'), '',
                          '',
                          single_core.get('AES-XTS'), single_core.get('Text Compression'),
                          single_core.get('Image Compression'), single_core.get('Navigation'), single_core.get('HTML5'),
                          single_core.get('SQLite'), single_core.get('PDF Rendering'),
                          single_core.get('Text Rendering'), single_core.get('Clang'), single_core.get('Camera'),
                          single_core.get('N-Body Physics'), single_core.get('Rigid Body Physics'),
                          single_core.get('Gaussian Blur'), single_core.get('Face Detection'),
                          single_core.get('Horizon Detection'), single_core.get('Image Inpainting'),
                          single_core.get('HDR'), single_core.get('Ray Tracing'),
                          single_core.get('Structure from Motion'), single_core.get('Speech Recognition'),
                          single_core.get('Machine Learning'),
                          multi_core.get('AES-XTS'), multi_core.get('Text Compression'),
                          multi_core.get('Image Compression'), multi_core.get('Navigation'), multi_core.get('HTML5'),
                          multi_core.get('SQLite'), multi_core.get('PDF Rendering'), multi_core.get('Text Rendering'),
                          multi_core.get('Clang'), multi_core.get('Camera'), multi_core.get('N-Body Physics'),
                          multi_core.get('Rigid Body Physics'), multi_core.get('Gaussian Blur'),
                          multi_core.get('Face Detection'), multi_core.get('Horizon Detection'),
                          multi_core.get('Image Inpainting'), multi_core.get('HDR'), multi_core.get('Ray Tracing'),
                          multi_core.get('Structure from Motion'), multi_core.get('Speech Recognition'),
                          multi_core.get('Machine Learning'), ben_info.get('Single-Core Score'),
                          ben_info.get('Multi-Core Score')]
        log.info('[Greekbech][collect_result]geekbench_list value:{}'.format(geekbench_list))

        geekbench_list1 = [0 if i is None else i for i in geekbench_list]
        log.info('[Greekbech][collect_result]geekbench_list1 value:{}'.format(geekbench_list1))
        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)

        sheet = data['Geekbench 5']
        max_row = sheet.max_row
        log.info("max row value is :{}".format(max_row))

        for i in range(len(geekbench_list1)):
            sheet.cell(max_row + 1, i + 1).value = geekbench_list1[i]

        data.save(self.save_path)
        data.close()
        time.sleep(3)
        log.info('[Greekbech][collect_result] collect_result ende')

        return True

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []
        for item in result:
            if abs(item.get('value') - average) / item.get('value') > deviation:
                abnormal_positions.append(item.get('position'))
                log.debug('[Geekbench][get abnormal data]Find abnormal data cell: {}'.format(item.get('position')))

        return abnormal_positions

    def test_performance(self):
        log.info('[geekbench][test_geekbench]begin to test')
        # for loop in range(self.__test_data.get('loop')):
        for loop in range(3):
            # self.launch()
            self.collect_result()
            log.info(f'[geekbench][test_geekbench]Test cycle {loop + 1} Finished')

    def analyze_report(self):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 10, row 1.
        :return:
        """
        global answer_num

        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)

        sheet = data['Geekbench 5']
        max_row = sheet.max_row
        max_col = sheet.max_column
        log.debug(
            '[Geekbench5][analyze_report]current used rows {}, start row {}'.format(max_row, self.data_start_col))

        for col in range(self.data_start_col, max_col + 1):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]

            for row in range(self.data_start_row, max_row + 1):
                cell_value = sheet.cell(row, col).value

                if cell_value == 'Average_Data':
                    continue

                cell_value = float(cell_value) if cell_value is not None else 0
                sum_data += cell_value
                temp_data = {'position': (row, col), 'value': cell_value}
                result.append(temp_data)

            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row + 1 - self.data_start_row)
                origin_num = Decimal(average_data)
                answer_num = origin_num.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

            sheet.cell(max_row + 1, col).value = answer_num
            sheet.cell(6, max_col + 1).value = 'Average_Data'.title()
            abnormals = self.__get_abnormal_data(result, average_data)

            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill

        data.save(self.save_path)
        data.close()


def start(value: dict):
    """
    main function for performance test
    1.TXT. support cross platform
    2. support all performance test items
    3. collect result and analyze test result
       3.1.TXT get result
       3.2 get average result
       3.3 mark abnormal data as red in excel
    4. write to excel as expected format
    5. all exception should be covered
    :return: None
    """
    import time
    perf = None
    try:

        if value.get('os') == 'wes':
            perf = Passmark()
        elif value.get('os') == 'linux':
            perf = Greekbech()
        perf.set_config(value)
        if perf.install():
            perf.install_license()
            perf.test_performance()
            perf.analyze_report()
            time.sleep(10)
        else:
            log.error('[start] install fail')

    except:
        import traceback
        log.error(f'[performance][start]get exception during test performance:{traceback.format_exc()}')
        traceback.print_exc()

    finally:
        log.info("[performance][start]end to test performance case")

    return None
