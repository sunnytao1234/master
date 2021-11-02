import os
import subprocess
import time

import openpyxl
from openpyxl.styles import PatternFill

from Common.common_function import OSTool
from Common.exception import IncorrectOS
from Common.log import log
from Common.common import get_lan_version, get_server_ip
from Test_Script.base.wireless import WlanWes, WlanLinux


class WlanTestLinux:

    def __init__(self):
        """
        router is interface for wlan to set value using different routers
        :param router:
        """
        self.router = None
        self.test_data = None

        self.save_path = OSTool.get_current_dir('result.xlsx')
        self.template_path = OSTool.get_current_dir('Test_Data', 'excle_template.xlsx')

        self.data_start_col = 8
        self.data_start_row = 3
        self.data_end_col = 20
        self.data_end_row = 6

    def enable_lan(self):  # enable lan device
        log.info("[WlanTestLinux][enable_lan]start enable eth0")
        os.popen('/usr/lib/hptc-network-mgr/common/netmgr_wired --up eth0')

        time.sleep(15)

    def disable_lan(self):
        """
        3
        """
        log.info("[WlanTestLinux][disable_lan]start disable eth0")
        os.popen('/usr/lib/hptc-network-mgr/common/netmgr_wired --down eth0')

        time.sleep(15)

    def set_config(self, value: dict):
        """
        1
        set wlan test config from analyzed data
        !! must call this method before test file read and write
        :param value: lan test config,
        eg. {'name': 'lan', 'platform': 'T630 DVT', 'loop': 3, 'status': None, 'tester': None, 'os': 'wes10 rs5'}
        :return: None
        """
        self.__test_data = value

        self.loop = value.get('loop')
        self.router = value.get('router')
        self.class_name = value.get('class_name')

        self.band = value.get('current_config').get('band')
        self.channel = value.get('current_config').get('channel')
        self.title = value.get('current_config').get('title')

        self.save_path = OSTool.get_current_dir('Test_Report',
                                                'Wlan_linux_{}_{}.xlsx'.format(value.get('platform'),
                                                                               value.get('os_info')))

        if self.class_name == 'AsuxAc87u' and self.band == '2.4g':
            self.ssid = 'AsuxAc87u'
            self.password = 'AsuxAc87u'
        elif self.class_name == 'AsuxAc87u' and self.band == '5g':
            self.ssid = 'AsuxAc87u_5G'
            self.password = 'AsuxAc87u'
        elif self.class_name == 'AsuxAx92u' and self.band == '5g-2':
            self.ssid = 'AsuxAx92u_5G-2'
            self.password = 'AsuxAx92u'
        elif self.class_name == 'AsuxAx92u' and self.band == '5g-1':
            self.ssid = 'AsuxAx92u_5G-1'
            self.password = 'AsuxAx92u'
        elif self.class_name == 'NetgearAx80' and self.band == '2.4g':
            self.ssid = 'NetgearAx80'
            self.password = 'NetgearAx80'
        self.linux = WlanLinux(ssid=self.ssid, password=self.password)

    def connect_to_ssid(self, check_cycle: int = 70, wait_time: int = 5) -> bool:
        """
        4
        connect to ap with given ssid name
        ap's ssid should exist in certain matrix
        :param ssid: ap's ssid name
        :return:
        """
        log.info(f"[wlanlinux][connect_to_ssid]connect to ssid")

        wlan_status = False
        if self.linux.disable_wlan():
            wlan_status = self.linux.enable_wlan()

        if wlan_status:
            for i in range(check_cycle):
                if self.linux.check_wlan_name():
                    log.info("[wlanlinux][connect_to_ssid]check wireless enable successfully")
                    break
                else:
                    time.sleep(wait_time)
                    continue

        self.linux.check_ssid_in_list()
        if not self.linux.check_ssid_in_list():
            return False

        self.linux.connect_wireless()
        if not self.linux.connect_wireless():
            return False

    def disconnect(self):
        """
        disconnect wlan
        :return:
        """

        self.linux.reset_env()

    def is_connected(self) -> bool:
        """
        5
        check whether wifi is connected
        :return: True if connected, else False
        """
        result = subprocess.getoutput('ping 192.168.1.1 -c 1')
        if '100% loss' in result:
            log.error("[wlan linux][access_internet]not access internet fail")
            return False
        else:
            log.info("[wlan linux][access_internet]access internet successful")
            return True

    def set_router_config(self, band, channel):
        """
        set router's band and channel value from admin url
        ! set value will restart router, wifi will disconnected
        :param value: value format from planned matrix
        :return:
        """
        pass

    def install(self):
        """
        2
        install iperf tool
        :return: boole
        """
        lib_version, jper_version = get_lan_version()
        tool_path = os.path.abspath(OSTool.get_current_dir("Test_Utility"))

        os.system('chmod 777 {}'.format(tool_path))

        log.info("[wlan linux][install]tool_path value:{}".format(tool_path))

        os.system('fsunlock')
        install_liperf = subprocess.getoutput('sudo dpkg -i {a}/{b}'.format(a=tool_path, b=lib_version))

        log.info("[wlan linux][install]install lib result is:{}".format(install_liperf))

        os.system('fsunlock')
        install_jper = subprocess.getoutput('sudo dpkg -i {a}/{b}'.format(a=tool_path, b=jper_version))

        log.info("[wlan linux][install] jperf result is:{}".format(install_jper))

        check_iperf3 = subprocess.getoutput('sudo dpkg -l | grep iperf3')

        log.info("[wlan-linux][install]check_iperf3 value:{}".format(check_iperf3))

        if check_iperf3 == '':
            log.error("[wlan-linux][install]install iperf3 fail")
            return False
        else:
            log.info("[wlan-linux][install]install iperf3 successful")
            return True

    def test_write(self):
        """
        test file transfer using jperf from server to local storage
        :return: average speed: Mb/s
        """
        test_write_three = []
        for i in range(self.loop):
            log.info("[wlan-linux][test_write]start to test lan")
            iperf3_run_path = subprocess.getoutput('which iperf3').strip().split('/iperf3')[0]
            log.info("[wlan-linux][test_write]iperf3_run_path value:{}".format(iperf3_run_path))
            cmd = " sudo /usr/bin/iperf3 -c {server_ip} -t 120 -w 56.0m".format(server_ip=get_server_ip()).replace('\n',
                                                                                                                   '')
            uut_linux_output = subprocess.getoutput(cmd)
            log.info("[wlan_linux][test_write]uut_linux_output value:{}".format(uut_linux_output))

            test_write_result = 0
            for line in uut_linux_output.split('\n')[::-1]:
                if 'Connection reset by peer' in line:
                    # Connection reset by peer
                    test_write_result = -1

                if 'Connection timed out' in line:
                    # Connection timed out
                    test_write_result = -2

                if 'receiver' in line:
                    test_write_result = float(line.split()[-3])

            log.info(f"[WlanTestWin][test_write]test_write_value is:{test_write_result}")

            test_write_three.append(round(test_write_result, 2))

            log.info(f'[WlanTestWin][test_write]end to test cycle[{i + 1}]')

        self.disconnect()
        time.sleep(30)
        return test_write_three

    def collect_result(self) -> None:
        """
        get test result and record to specific excel file
        :return:
        """
        log.info(f"[WlanTestWin][collect_result]start to collect results")

        test_write_value = self.test_write()
        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path, )
        else:
            data = openpyxl.load_workbook(self.template_path)
        current_col = 0
        if self.title == '2.4g/CH-3/BW-20/802.11n':
            current_col = 8
        elif self.title == '2.4g/CH-11/BW-40/802.11n':
            current_col = 9
        elif self.title == '2.4g/CH-5/BW-20/40/802.11ax':
            current_col = 10
        elif self.title == '5g/CH-153/BW-20/802.11n':
            current_col = 11
        elif self.title == '5g/CH-157/BW-40/802.11n':
            current_col = 12
        elif self.title == '5g/CH-153/BW-20/802.11ac':
            current_col = 13
        elif self.title == '5g/CH-157/BW-40/802.11ac':
            current_col = 14
        elif self.title == '5g/CH-157/BW-80/802.11ac':
            current_col = 15
        elif self.title == '5g-2/CH-128/BW-160/802.11ax':
            current_col = 19
        elif self.title == '5g-2/CH-157/BW-80/802.11ax':
            current_col = 18
        elif self.title == '5g-2/CH-157/BW-40/802.11ax':
            current_col = 17
        elif self.title == '5g-2/CH-157/BW-20/802.11ax':
            current_col = 16
        log.info('[wlanwes][collect_result] current_col value:{}'.format(current_col))
        lan_list = ['', '', '', self.__test_data.get('platform_info'),
                    self.__test_data.get('bios_info'), self.__test_data.get('ml_info'),
                    self.__test_data.get('os_info'), '']

        sheet = data['WLAN']

        for i in range(self.loop):
            max_row = 2
            log.info("[WlanTestWin][collect_result]max row value is :{}".format(max_row))
            lan_list[-1] = test_write_value[i]
            for m in range(len(lan_list)):
                if m + 1 > 0 and m + 1 < 8:
                    sheet.cell(max_row + i + 1, m + 1).value = lan_list[m]
                else:
                    sheet.cell(max_row + i + 1, current_col).value = lan_list[-1]
            data.save(self.save_path)
            data.close()

        log.info(f"[WlanTestWin][collect_result]end to collect results")

    def test_wlan(self):
        """
        6
          run using specific config
          :return:
        """
        self.collect_result()

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []
        for item in result:
            if item.get('value') == 0:
                abs_value = abs(item.get('value') - average)
            else:
                abs_value = abs((item.get('value') - average) / item.get('value'))
            if abs_value > deviation:
                abnormal_positions.append(item.get('position'))
                log.debug('[wlan][get abnormal data]Find abnormal data cell: {}'.format(item.get('position')))
        return abnormal_positions

    def analyze_report(self):
        """
        7
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 7, row 2
        :return:
        """
        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)
        sheet = data['WLAN']
        max_row = self.data_end_row
        for col in range(self.data_start_col, self.data_end_col):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
            for row in range(self.data_start_row, max_row):
                cell_value = sheet.cell(row, col).value
                cell_value = float(cell_value) if cell_value is not None else 0
                sum_data += cell_value
                temp_data = {'position': (row + 1, col), 'value': cell_value}
                result.append(temp_data)
            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row - self.data_start_row)
            sheet.cell(max_row, col).value = average_data
            sheet.cell(6, 20).value = 'Avage_data'.title()
            abnormals = self.__get_abnormal_data(result, average_data)
            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill
        data.save(self.save_path)
        data.close()


class WlanTestWin:

    def __init__(self):
        """
        router is interface for wlan to set value using different routers
        :param router:
        """
        self.router = None
        self.test_data = None

        self.save_path = OSTool.get_current_dir('result.xlsx')
        self.template_path = OSTool.get_current_dir('Test_Data', 'excle_template.xlsx')

        self.data_start_col = 8
        self.data_start_row = 3
        self.data_end_col = 20
        self.data_end_row = 6

        self.global_path = OSTool.get_current_dir("Test_Data", 'td_common', 'global_config.yml')

    def enable_lan(self):
        log.info("[WlanTestWin][enable]enable lan")
        self.wes.enable_lan()
        log.info("[WlanTestWin][enable]enable lan successfully")
        # enable wlan device

    def disable_lan(self):
        log.info("[WlanTestWin][disable]disable lan")
        self.wes.disable_lan()
        log.info("[WlanTestWin][disable]disable lan successfully")

    def connect_to_ssid(self):
        """
        connect to ap with given ssid name
        ap's ssid should exist in certain matrix
        :param ssid: ap's ssid name
        :return:
        """
        log.info("[WlanTestWin][connect_to_ssid]connect to ssid")
        # self.wes.change_wifi_switch(action='open')
        self.wes.connect_wifi()
        log.info("[WlanTestWin][connect_to_ssid]connect to ssid successfully")

    def disconnect(self) -> bool:
        """
        disconnect wlan
        :return:
        """
        log.info("[WlanTestWin][disconnect]disconnect to ssid")
        return self.wes.reset_env()

    def is_connected(self) -> bool:
        """
        check whether wifi is connected
        :return: True if connected, else False
        """
        result = subprocess.getoutput('ping 192.168.1.1 -c 1')
        log.info(f"ping 192.168.1.1 result:{result}")

        if '100% loss' in result:
            log.error("[WlanTestWin][is_connected]access internet unsuccessfully")

            return False
        else:
            log.info("[WlanTestWin][is_connected]access internet successfully")

            return True

    def set_config(self, value: dict):
        """
           set wlan test config from analyzed data
           !! must call this method before test file read and write
           :param value: lan test config,
           eg. {'name': 'lan', 'platform': 'T630 DVT', 'loop': 3, 'status': None, 'tester': None, 'os': 'wes10 rs5'}
           :return: None
           """
        self.__test_data = value

        log.info(f"[WlanTestWin][set_config]get value from remote:{value}")

        self.loop = value.get('loop')
        self.class_name = value.get('class_name')

        self.band = value.get('current_config').get('band')
        self.channel = value.get('current_config').get('channel')
        self.title = value.get('current_config').get('title')

        self.save_path = OSTool.get_current_dir('Test_Report',
                                                'Wlan_wes_{}_{}.xlsx'.format(value.get('platform'),
                                                                             value.get('os_info')))

        if self.class_name == 'AsuxAc87u' and self.band == '2.4g':
            self.ssid = 'AsuxAc87u'
        elif self.class_name == 'AsuxAc87u' and self.band == '5g':
            self.ssid = 'AsuxAc87u_5G'
        elif self.class_name == 'AsuxAx92u' and self.band == '5g-2':
            self.ssid = 'AsuxAx92u_5G-2'
        elif self.class_name == 'AsuxAx92u' and self.band == '5g-1':
            self.ssid = 'AsuxAx92u_5G-1'
        elif self.class_name == 'NetgearAx80' and self.band == '2.4g':
            self.ssid = 'NetgearAx80'
        else:
            log.warning(f"current argument is not supported:[{self.class_name}/{self.band}]")

        log.info(f"[WlanTestWin][set_config] current ssid to connect is:{self.ssid}")

        self.wes = WlanWes(ssid=self.ssid)

    def set_router_config(self, band, channel):
        """
        set router's band and channel value from admin url
        ! set value will restart router, wifi will disconnected
        :param value: value format from planned matrix
        :return:
        """
        pass

    def install(self):
        """
        install iperf tool
        :return: boole
        wes not need install iperf3
        """
        pass

    def test_write(self):
        """
        test file transfer using jperf from local storage to server
        :return: average speed: Mb/s
        """
        test_write_three = []

        log.info('[WlanTestWin][test_write]begin to test')

        for loop in range(self.loop):
            log.info(f'[WlanTestWin][test_write]begin to test cycle[{loop + 1}/{self.loop}]')

            version = get_lan_version()
            uut_iperf_path = OSTool.get_current_dir('Test_Utility', '{}'.format(version))

            cmd = "{tool} -c {server_ip} -t 120 ".format(tool=uut_iperf_path, server_ip=get_server_ip())
            log.info(f"[WlanTestWin][test_write]run cmd:{uut_iperf_path} -c {get_server_ip()} -t 120")

            uut_wes_output = subprocess.getoutput(cmd)
            log.info("[WlanTestWin][test_write]uut_output value is:\n{}\n".format(uut_wes_output))

            test_write_result = 0
            for line in uut_wes_output.split('\n')[::-1]:
                if 'Connection reset by peer' in line:
                    # Connection reset by peer
                    test_write_result = -1

                if 'Connection timed out' in line:
                    # Connection timed out
                    test_write_result = -2

                if 'receiver' in line:
                    test_write_result = float(line.split()[-3])

            log.info(f"[WlanTestWin][test_write]test_write_value is:{test_write_result}")

            test_write_three.append(round(test_write_result, 2))

            log.info(f'[WlanTestWin][test_write]end to test cycle[{loop + 1}]')


        self.disconnect()

        log.info(f"[WlanTestWin][test_write]end to test:<{self.router},{self.ssid}>")

        return test_write_three

    def collect_result(self) -> None:
        """
        get test result and record to specific excel file
        :return:
        """
        log.info(f"[WlanTestWin][collect_result]start to collect results")

        test_write_value = self.test_write()
        log.debug('[WlanTestWin][collect_result] test_write_value:{}'.format(test_write_value))
        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)
        current_col = 1
        log.info('[WlanTestWin][collect_result] self.title value:{}'.format(self.title))
        if self.title == '2.4g/CH-3/BW-20/802.11n':
            current_col = 8
        elif self.title == '2.4g/CH-11/BW-40/802.11n':
            current_col = 9
        elif self.title == '2.4g/CH-5/BW-20/40/802.11ax':
            current_col = 10
        elif self.title == '5g/CH-153/BW-20/802.11n':
            current_col = 11
        elif self.title == '5g/CH-157/BW-40/802.11n':
            current_col = 12
        elif self.title == '5g/CH-153/BW-20/802.11ac':
            current_col = 13
        elif self.title == '5g/CH-157/BW-40/802.11ac':
            current_col = 14
        elif self.title == '5g/CH-157/BW-80/802.11ac':
            current_col = 15
        elif self.title == '5g-2/CH-128/BW-160/802.11ax':
            current_col = 19
        elif self.title == '5g-2/CH-157/BW-80/802.11ax':
            current_col = 18
        elif self.title == '5g-2/CH-157/BW-40/802.11ax':
            current_col = 17
        elif self.title == '5g-2/CH-157/BW-20/802.11ax':
            current_col = 16

        log.info('[wlanwes][collect_result] current_col value:{}'.format(current_col))
        lan_list = ['', '', '', self.__test_data.get('platform_info'),
                    self.__test_data.get('bios_info'), self.__test_data.get('ml_info'),
                    self.__test_data.get('os_info'), '']
        log.debug('[wlan_wes][collect_result] lan list value:{}'.format(lan_list))
        sheet = data['WLAN']
        max_row = 2
        log.info("[WlanTestWin][collect_result]max row value is :{}".format(max_row))

        for i in range(self.loop):
            lan_list[-1] = test_write_value[i]
            for m in range(len(lan_list)):
                if m + 1 > 0 and m + 1 < 8:
                    sheet.cell(max_row + i + 1, m + 1).value = lan_list[m]
                else:
                    sheet.cell(max_row + i + 1, current_col).value = lan_list[-1]

            data.save(self.save_path)
            data.close()

        log.info(f"[WlanTestWin][collect_result]end to collect results")

    def test_wlan(self):
        """
        6
          run using specific config
          :return:
          need  Need to be perfected
        """
        log.info("[WlanTestWin][test_wlan]start to test wlan")

        self.collect_result()

        log.info("[WlanTestWin][test_wlan]end to test wlan")

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """t
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []

        for item in result:
            if item.get('value') == 0:
                abs_value = abs(item.get('value') - average)
            else:
                abs_value = abs((item.get('value') - average) / item.get('value'))

            if abs_value > deviation:
                abnormal_positions.append(item.get('position'))

                log.debug('[WlanTestWin][get_abnormal_data]find abnormal data cell: {}'.format(item.get('position')))

        return abnormal_positions

    def analyze_report(self):
        """
        7
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 8, row 3
        :return:
        """
        log.info("[WlanTestWin][analyze_report]start to analyze report")

        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)
        sheet = data['WLAN']
        max_row = self.data_end_row
        log.info('[WlanTestWin][analyze_report]max_row value:{}'.format(max_row))

        for col in range(self.data_start_col, self.data_end_col):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]

            for row in range(self.data_start_row, max_row):
                cell_value = sheet.cell(row, col).value
                cell_value = round(float(cell_value), 2) if cell_value is not None else 0

                sum_data += cell_value
                temp_data = {'position': (row + 1, col), 'value': cell_value}
                result.append(temp_data)

            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row - self.data_start_row)

            sheet.cell(max_row, col).value = round(average_data, 2)
            sheet.cell(6, 20).value = 'Average_data'.title()

            abnormals = self.__get_abnormal_data(result, average_data)

            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill

        log.info("[WlanTestWin][analyze_report]end to analyze report")

        data.save(self.save_path)
        data.close()


def start(value: dict):
    """"""
    log.info("[wlan][start]begin to test wlan case")

    try:
        if value.get('os') == 'wes':
            wlan = WlanTestWin()
        elif value.get('os') == 'linux':
            wlan = WlanTestLinux()
        else:
            log.error('[wlan][start]Incorrect os given')
            raise IncorrectOS('Incorrect OS given')

        wlan.set_config(value)

        wlan.connect_to_ssid()

        time.sleep(15)

        if wlan.is_connected():
            wlan.install()
            wlan.test_wlan()
            wlan.analyze_report()
            time.sleep(15)
            wlan.is_connected()
        else:
            log.error(f"[wlan][start]cannot connect to ssid")


    except:
        import traceback
        log.error(f'[wlan][start]get exception during test wlan:{traceback.format_exc()}')
        traceback.print_exc()

    finally:
        log.info("[wlan][start]end to test wlan case")

    return None
