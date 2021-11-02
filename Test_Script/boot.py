# -*- coding: utf-8 -*-
# @time     :   4/1/2021 3:35 PM
# @author   :   balance
# @file     :   boot.py
# todo:develop and debug
import os
import statistics
import time
import typing

import openpyxl
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet

from Common.common_function import get_current_dir
from Common.log import log
from Test_Script.base import socket_action
from Test_Script.base.bootstrap import BaseBootTest, BootResult

log.log_level = 4


class WesBootTest(BaseBootTest):
    def __init__(self, config: dict):
        super(WesBootTest, self).__init__(config=config,
                                          platform=config['platform'],
                                          boot_wait_time=config['boot_wait_time'],
                                          camera_device_index=config['camera_device_index'],
                                          manual_boot_list=config['manual_boot_list'])

    def test_result(self) -> typing.List[BootResult]:
        log.debug("[WesBootTest][test_result]the test_result process is started")

        results = []

        for cycle in range(self.config['loop']):
            log.info(f"[WesBootTest][test_result]current test progress: [{cycle + 1}/{self.config['loop']}]")

            if self.observer.power_on(self.config['platform'], uut_mac=self.config['mac_info'],
                                      uut_broadcast=self.config['uut_broadcast']):
                self.timer.set_time_report(startup_time=time.time())
                if self.observer.check_hp_logo():
                    self.timer.set_time_report(hp_logo_time=time.time())
                    if self.observer.check_desktop():
                        self.timer.set_time_report(desktop_time=time.time())

                        time.sleep(45)
                        if cycle == self.config['loop'] - 1:
                            self.observer.reboot()
                        else:
                            if self.observer.power_off():
                                log.info(f"[WesBootTest][test_result]power off uut")
                            else:
                                log.error(f"[WesBootTest][test_result]cannot power off")
                    else:
                        log.error(f"[WesBootTest][test_result]desktop does not appear")
                else:
                    log.error(f"[WesBootTest][test_result]hp logo does not appear")
            else:
                log.error(f"[WesBootTest][test_result]cannot power on")

            results.append(self.timer.get_time_report())

            log.info(f"[WesBootTest][test_result]please wait to {self.boot_wait_time}s")
            time.sleep(self.boot_wait_time)
        log.debug("[WesBootTest][test_result]the test_result process is ended")

        return results

    def analyse_result(self, results: typing.List[BootResult]) -> typing.Tuple[
        typing.List, typing.List, typing.List, typing.List]:
        log.debug("[WesBootTest][analyse_result]the analyse_result process is started")

        bios_post_times, os_startup_times, total_times, average_times = [], [], [], []

        for index, row_data in enumerate(results):
            bios_post_times.append(row_data.hp_logo_time - row_data.startup_time)
            os_startup_times.append(row_data.desktop_time - row_data.hp_logo_time)
            total_times.append(row_data.desktop_time - row_data.startup_time)

        average_times.append(statistics.mean(bios_post_times))
        average_times.append(statistics.mean(os_startup_times))
        average_times.append(statistics.mean(total_times))

        log.debug("[WesBootTest][analyse_result]the analyse_result process is ended")

        return bios_post_times, os_startup_times, total_times, average_times

    def write_result(self, results: typing.Tuple[typing.List, typing.List, typing.List, typing.List],
                     template_path: str = get_current_dir('Test_Data', 'excle_template.xlsx'),
                     fill_color=styles.PatternFill("solid", fgColor="FF0000"),
                     abnormal_threshold: float = 0.05
                     ):
        log.debug("[WesBootTest][write_result]the write_result process is started")

        save_path = get_current_dir('Test_Report',
                                    'Boot_{}_{}.xlsx'.format(self.config['platform'], self.config['os']))

        if os.path.exists(save_path):
            excel = openpyxl.load_workbook(save_path)
        else:
            excel = openpyxl.load_workbook(template_path)

        try:
            base_info = [
                '', self.config['disk_info'], self.config['disk_info'].split()[-1], '',
                self.config['platform_info'], self.config['memory_info'], self.config['cpu_info'],
                self.config['gpu_info'],
                '', self.config['bios_info'], self.config['os_info'] + self.config['ml_info'], self.config['ml_info'],
            ]
        except Exception:
            base_info = [
                '', self.config['disk_info'], '', '',
                self.config['platform_info'], self.config['memory_info'], self.config['cpu_info'],
                self.config['gpu_info'],
                '', self.config['bios_info'], self.config['os_info'] + self.config['ml_info'], self.config['ml_info'],
            ]

        boot_sheet: Worksheet = excel['Booting Time ']
        current_max_row, current_max_col = boot_sheet.max_row, boot_sheet.max_column

        boot_sysinfo_cells = boot_sheet['A{}'.format(current_max_row + 1):'L{}'.format(current_max_row + 4)]
        log.info(
            f"[WesBootTest][write_result]current boot system info cells: "
            f"({'A{}'.format(current_max_row + 1)}:{'L{}'.format(current_max_row + 4)})")

        boot_result_cells = boot_sheet['M{}'.format(current_max_row + 1):'O{}'.format(current_max_row + 4)]
        log.info(
            f"[WesBootTest][write_result]current boot time result cells: "
            f"({'M{}'.format(current_max_row + 1)}:{'O{}'.format(current_max_row + 4)})")

        for boot_sysinfo_cell_row in boot_sysinfo_cells:
            for index, cell in enumerate(boot_sysinfo_cell_row):
                cell.value = base_info[index]
                log.info(f"[WesBootTest][write_result]current base info is: {base_info[index]}")

        for row, boot_result_cell_row in enumerate(boot_result_cells):
            # write average data to excel
            if row == len(boot_result_cells) - 1:
                for col, cell in enumerate(boot_result_cell_row):
                    cell.value = round(results[row][col], 2)
                    log.info(f"[WesBootTest][write_result]current cell average value is: {cell}")

            # write epoch data to excel
            else:
                for col, cell in enumerate(boot_result_cell_row):
                    cell.value = round(results[col][row], 2)
                    log.info(f"[WesBootTest][write_result]current cell value is: {cell}")

                    average_value = round(results[-1][row], 2)
                    log.info(f"[WesBootTest][write_result]current average value is: {average_value}")

                    if average_value == 0:
                        cell.fill = fill_color
                        log.info(f"[WesBootTest][write_result]abnormal value is founded:{cell}")

                    elif abs(cell.value - average_value) / average_value > abnormal_threshold:
                        cell.fill = fill_color
                        log.info(f"[WesBootTest][write_result]abnormal value is funded:{cell}")

        excel.save(save_path)
        log.info(f"[WesBootTest][write_result]result excel save to:{save_path}")

        log.debug("[WesBootTest][write_result]the write_result process is ended")


def start(value: dict):
    log.debug("[boot][start]the boot test process is started")

    if value['os'] not in ('wes', 'linux'):
        log.error(f'[boot][start]a incorrect os argument is given:{value["os"]}')
    else:
        with WesBootTest(config=value) as boot:
            result = boot.test_result()
            result = boot.analyse_result(result)
            boot.write_result(result)

            log.info(f"[boot][start]current boot-time result: {result}")

    log.debug("[boot][start]the boot test process is ended")
