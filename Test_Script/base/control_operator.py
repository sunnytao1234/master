from Common.log import log
import time
from Test_Script.base.common import OSTool, read_function_tool_yaml
import os
import re
from datetime import datetime
from Common.picture_operator import ComparePic

from Common.camera import CAMERA
from Common.man import Man
from Common.valve import ValveContainer
from Test_Script.base.common import wol_wakeup, get_serial_port,recognize_string
import openpyxl
from openpyxl.styles import PatternFill
import zipfile

class control_operator:

    def __init__(self):
        self.value = ValveContainer(port=get_serial_port('CH34'))
        self.camera = CAMERA()
        self.man = Man()
    def control_recode_capture_start_time(self):
        """
        control run
        check logo appear
        write time on control txt
        :return:
        """
        start_time = datetime.now()
        start_txt_path = OSTool.get_current_dir("Test_Report",'capture_start.txt')
        if os.path.exists(start_txt_path):
            os.remove(start_txt_path)
        with open(start_txt_path,'w') as recode:
            recode.write('{}'.format(start_time))
        return start_time



    def control_check_capture_no_sign(self):
        """
        control run
        :return:
        """
        com = ComparePic()
        screen_path = OSTool.get_current_dir('Test_Report', 'capture.png')
        save_path = OSTool.get_current_dir('Test_Data', 'td_thinupdate', 'capture', 'capture_shutdown')
        result = com.compare_picture(file=screen_path, template_folder=save_path)
        if result:
            log.info('[control_operator][control_check_capture_no_sign] uut is shut down')
            return True
        else:
            return False


    def control_main_capture_process(self,platform, check_cycle: int = 3600, wait_time: int = 1):
        """control run"""
        capture_apper_picture = 0
        for i in range(check_cycle):
            capture_picture = OSTool.get_current_dir('Test_Report','capture.png')
            self.camera.screenshot(file=capture_picture)
            ocr_capture_value = str('_').join(recognize_string(capture_picture, platform_name=platform, lang='eng'))
            log.info('[control_operator][main_process] ocr value :{}'.format(ocr_capture_value))
            result = self.control_check_capture_no_sign()
            log.info('[control_operator][control_main_capture_process] result value:{}'.format(result))
            if result:
                    capture_apper_picture = capture_apper_picture + 1
                    log.info(f'[control_operator][deploy_result] capture appear picture value:{capture_apper_picture}')
                    if capture_apper_picture == 1:
                        time.sleep(150)
                    elif capture_apper_picture == 2:
                        # log.info('[control_operator][main_capture_process] capture is complete')
                        self.read_deploy_time(txtnsme='capture_ended.txt')
                        return True
            elif 'error' in recognize_string(capture_picture,platform_name=platform, lang='eng'):
                log.info('[control_operator][main_capture] capture fail')
                return False
            else:
                if os.path.exists(capture_picture):
                    os.remove(capture_picture)
                    time.sleep(wait_time)
                    continue

        log.error('[control_operator][control_main_capture_process]not found error and 100# time out ')
        return False


    def control_recover_env(self):
        """:
        control run
        if capture appear error,
        you need recover env
        """
        self.man.input_string('exit')
        self.man.wait(1)
        self.man.press_key('Enter')


    def control_open_value(self):
        self.value.open()

    def control_close_value(self):
        self.value.close()

    def read_deploy_time(self,txtnsme):
        recode_time = datetime.now()
        start_txt_path = OSTool.get_current_dir("Test_Report", txtnsme)
        if os.path.exists(start_txt_path):
            os.remove(start_txt_path)
        with open(start_txt_path, 'w') as recode:
            recode.write('{}'.format(recode_time))
        return recode_time

    def control_difference_time(self,start_time_name,end_time_name):
        with open('{}'.format(OSTool.get_current_dir("Test_Report", end_time_name)), 'r') as f:
            end_time = f.read()
            end_list = end_time.split(' ')[1].split(':')

            log.info('[control_calculate_time]end list value:{}'.format(end_list))
            end_time_total = int(end_list[0]) * 3600 + int(end_list[1]) * 60 + int(end_list[2].split('.')[0])
            log.info('[control_calculate_time]end end_time_total value:{}'.format(end_time_total))
        with open('{}'.format(OSTool.get_current_dir("Test_Report", start_time_name)), 'r') as f1:
            start_time = f1.read()
            start_list = start_time.split(' ')[1].split(':')
            log.info('[control_calculate_time]end list value:{}'.format(start_list))
            start_time_total = int(start_list[0]) * 3600 + int(start_list[1]) * 60 + int(start_list[2].split('.')[0])
            log.info('[control_calculate_time]end start_time_total value:{}'.format(start_time_total))
        calculate_time_value = int(end_time_total) - int(start_time_total)
        if calculate_time_value > 3600:
            hour = int((calculate_time_value / 3600))
            print(hour)
            min = int((calculate_time_value / 60))
            log.info('min value:{}'.format(min))
            second = calculate_time_value - min * 60
            log.info('second value:{}'.format(second))
        else:
            hour = 0
            print(hour)
            min = int(calculate_time_value / 60)
            log.info('min value：{}'.format(min))
            second = calculate_time_value - min * 60
            log.info('second value:{}'.format(second))
        time_value = '{min}:{second}'.format(min=min, second=second)
        log.info('[control_calculate_time]calculate_time_value:{}'.format(calculate_time_value))
        return time_value

    def control_deploy_process(self, loop,platform, check_cycle: int = 3600, wait_time: int = 1):
        self.man.wait(200)
        self.man.input_string("Y")
        self.man.press_key('Enter')
        self.man.wait(2)
        self.man.input_string("Y")
        self.man.press_key('Enter')
        self.read_deploy_time(txtnsme='deploy_start_time.txt')
        for i in range(check_cycle):
            deploy_pic = OSTool.get_current_dir('Test_Report', 'deploy.png')
            deploy_new_name = OSTool.get_current_dir('Test_Report', 'deploy_{}.png'.format(loop))
            if os.path.exists(deploy_new_name):
                os.remove(deploy_new_name)
            self.camera.screenshot(file=deploy_pic)
            deploy_ocr_value = str('_').join(recognize_string(deploy_pic,platform_name=platform, lang='eng'))
            log.info(f'deploy_ocr value:{deploy_ocr_value}')
            deploy_value_list = ['successfully','succousfully','sacessfalgt', 'sacessfalyt', 'sacossfalyt', 'sacessfagt',
                                 'sacossfagt', 'sacossfalgt']
            for m in deploy_value_list:
                if m in deploy_ocr_value:
                    if os.path.exists(deploy_pic):
                        os.rename(deploy_pic, deploy_new_name)

                    self.read_deploy_time(txtnsme='deploy_ended_time.txt')

                    deploy_ocr = re.findall('\d{1,2}:\d{1,2}', deploy_ocr_value)
                    log.info('deploy_ocr value:{}'.format(deploy_ocr))
                    if len(deploy_ocr)>0:
                        if ':' in deploy_ocr[0] and len(deploy_ocr) == 1:
                            deploy_ocr_result = deploy_ocr[0]
                            self.man.wait(30)
                            self.man.input_string("S")
                            self.man.wait(60)
                            return deploy_ocr_result

                    else:
                        deploy_ocr = re.findall('\d{4}', deploy_ocr_value)
                        if len(deploy_ocr)>0:
                            value = list(deploy_ocr[0])
                            alter_value = value.insert(2, ':')
                            log.info('value is:{}'.format(value))
                            deploy_ocr_result = str().join(value)
                            log.info('deploy_ocr_result value:{}'.format(deploy_ocr_result))
                            self.man.wait(30)
                            self.man.input_string("S")
                            self.man.wait(60)
                            return deploy_ocr_result

                elif 'error' in deploy_ocr_value:
                    log.error('[control_operator][control_deploy_process] deploy appear not recover error')
                    return False
                else:
                    time.sleep(wait_time)
                    continue


    def control_wakeup_uut(self, platform, mac):
        platform_list = ['mt32','mt46']
        if platform in platform_list:
            log.info('[control_operator][control_wakeup_uut] not wakeup uut online please docket')
            wol_wakeup(mac)

        else:
            log.info('[control_operator][control_wakeup_uut] wakeup uut online')
            wol_wakeup(mac)
        return True


    def control_wait_time(self, wait_time: int= 60) -> object:
        time.sleep(wait_time)
        return wait_time


    def control_switch_user(self, user='Admin'):
        for i in range(3):
            self.man.press_key_delay('Ctrl+Alt+Del', 5)
        self.man.press_key_delay('Down', 1)
        self.man.press_key_delay('Enter', 30)
        self.man.press_key_delay('Enter', 10)
        self.man.press_key('Tab', 3)
        self.man.wait(1)
        self.man.press_key('Enter')
        self.man.input_string(user)
        self.man.press_key_delay('Enter', 80)


    def __get_abnormal_data(self, result: list, avage_data_value: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []

        for item in result:
            if item.get('value') == 0:
                abs_value = abs(item.get('value') - avage_data_value)
            else:
                abs_value = abs((item.get('value') - avage_data_value) / item.get('value'))

            if abs_value > deviation:
                abnormal_positions.append(item.get('position'))

                log.debug('[Capture][get_abnormal_data]find abnormal data cell: {}'.format(item.get('position')))

        return abnormal_positions

    def analyze_report(self, save_path='',):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 10, row 1
        :return:
        """
        data_start_col = 9
        data_start_row = 2
        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(save_path)
        sheet = data['Thinupdate']
        max_row = sheet.max_row
        print('max_row value:{}'.format(max_row))
        log.debug(
            '[Thinupdate][analyze_report]current used rows {}, start row {}'.format(max_row, data_start_row))
        for col in range(data_start_col, data_start_col + 3):
            sum_data = 0
            result = []
            for row in range(data_start_row, max_row + 1):
                cell_value = sheet.cell(row, col).value
                if ':' in str(cell_value):
                    cell_value_list = str(cell_value).split(':')
                    value = int(cell_value_list[0]) * 60 + int(cell_value_list[1])
                    sum_data += value
                    temp_data = {'position': (row, col), 'value': value}
                    result.append(temp_data)
                    print('result value: {}'.format(result))
            if sum_data == 0:
                avage_data_value = 0
                average_data = 0
            else:
                avage_data_value = sum_data / 3
                print('avage_data_value :{}'.format(avage_data_value))
                if avage_data_value <= 59:
                    average_data = '00:{}'.format(int(str(avage_data_value).split('.')[0]))
                else:
                    second = avage_data_value % 60
                    min = (avage_data_value - second) / 60
                    average_data = '{a}:{b}'.format(a=int(str(min).split('.')[0]), b=int(str(second).split('.')[0]))

            sheet.cell(max_row + 1, col).value = average_data
            abnormals = self.__get_abnormal_data(result, avage_data_value)
            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill
        data.save(save_path)
        data.close()

    def zip_dir(self, skip_img=True,zipname ='function Automation'):
        """
        zip specific file
        :param skip_img: skip img folder in Test_Report
        :param report_name:
        :return:
        """
        platform_info = zipname

        filename = OSTool.get_current_dir('{}.zip').format(platform_info)
        filename = filename.replace(' ', '_')

        zip_file = zipfile.ZipFile(filename, "w", zipfile.ZIP_DEFLATED)
        for path, dirnames, filenames in os.walk(OSTool.get_current_dir('Test_Report')):
            # 去掉目标跟路径，只对目标文件夹下边的文件及文件夹进行压缩
            fpath = path.replace(OSTool.get_current_dir('Test_Report'), '')
            if 'img' in dirnames:
                if skip_img:
                    dirnames.remove('img')
            for name in filenames:
                zip_file.write(os.path.join(path, name), os.path.join(fpath, name))
        zip_file.close()
        return filename