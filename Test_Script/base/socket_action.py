import json
import os
import socket
import threading
import time
import traceback
import typing
from queue import Queue

import openpyxl

from Common import email_tool
from Common.common_function import OS
from Common.log import log
from Test_Script import wlan, thin
from Test_Script.base import bootstrap
from Test_Script.base.common import get_server_ip, OSTool
from Test_Script.base.common import is_controller
from Test_Script.base.common import read_uut_info
from Test_Script.base.control_operator import control_operator
from Test_Script.base.uut_operator import uut_operator

if OS == 'Windows':
    import winsound
    from Common.tc_info import WESInfo

    wes = WESInfo()

# must be global variable
_REQ_MSG_QUEUE = Queue()
_RES_MSG_QUEUE = Queue()

if is_controller():
    control = control_operator()
uut = uut_operator()

if is_controller():
    uut_info_list = read_uut_info()


def get_uut_ip():
    with open('uut.txt', 'r') as f:
        uut_ip = f.read()
        return uut_ip


class SocketAgent:
    def __init__(self, server_name: str = 'socket_server',
                 listen_ip: str = '0.0.0.0', listen_port: int = 2333,
                 queue_maxsize: int = 0
                 ):
        self._server_name = server_name

        self._socket = None

        self._listen_ip = listen_ip
        self._listen_port = listen_port

        # put and get messages from client
        self._req_msg_queue = _REQ_MSG_QUEUE
        self._res_msg_queue = _RES_MSG_QUEUE
        self._req_msg_queue.maxsize = queue_maxsize
        self._res_msg_queue.maxsize = queue_maxsize

        # put and get exceptions from child thread
        self.exception_queue = Queue(maxsize=queue_maxsize)

        self._socket_server = None

        log.info(
            f"[SocketAgent][init]SocketAgent creation succeed:\n"
            f"listen ip:{self._listen_ip},listen port:{self._listen_port}\n")

    @property
    def res_queue(self):
        return self._res_msg_queue

    @property
    def req_queue(self):
        return self._req_msg_queue

    @classmethod
    def create_socket_agent(cls, server_name: str = 'create_server',
                            listen_ip: str = '0.0.0.0', listen_port: int = 2333,
                            queue_maxsize: int = 0, clear_queue: bool = True) -> typing.Any:
        agent: SocketAgent = cls(server_name=server_name,
                                 listen_ip=listen_ip, listen_port=listen_port,
                                 queue_maxsize=queue_maxsize)
        if clear_queue:
            while not agent.req_queue.empty():
                agent.req_queue.get()
            # while not agent.res_queue.empty():
            #     agent.res_queue.get()
            while not agent.exception_queue.empty():
                agent.exception_queue.get()

            log.info("[SocketAgent][create_socket_agent]exception queue was cleared")

        return agent

    def open_server(self, daemon: bool = False, bufsize: int = 4096,
                    backlog: int = 5, keepalive: bool = False,
                    blocking: bool = True, timeout_value: float = 0.0) -> threading.Thread or None:
        server = None

        try:
            self._socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self._socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            self._socket.setblocking(blocking)

            if float(timeout_value) != 0.0:
                self._socket.settimeout(timeout_value)

            log.info(f"[SocketAgent][open_server]current socket config:{blocking},{timeout_value}")

            self._socket_server = threading.Thread(name=self._server_name, target=self._run_socket,
                                                   kwargs={'backlog': backlog, 'keepalive': keepalive,
                                                           'bufsize': bufsize})
            self._socket_server.setDaemon(daemon)
            self._socket_server.start()

            server = self._socket_server
        except Exception as e:
            log.error(f"[SocketAgent][open_server]raise exception by creating server:{traceback.format_exc()}")
            server = None
        finally:
            return server

    def close_server(self, msg: str = 'close', block: bool = True, timeout: int = 3) -> bool:
        status = False

        if not self.req_queue.full():
            socket.create_connection(('127.0.0.1', self._listen_port))
            self.req_queue.put(msg, block=block, timeout=timeout)
            time.sleep(timeout)

            if not self._socket_server.is_alive():
                log.info(
                    f"[SocketAgent][close_server]current socket server thread status:{self._socket_server.is_alive()}")

                self._socket = None
                self._socket_server = None

                status = True

                log.info(f"[SocketAgent][close_server]close server successfully")
        else:
            log.error(f"[SocketAgent][close_server]socket request message queue is fulled")

        return status

    def get_server_info(self) -> dict:
        server_info = {
            'server_name': self._server_name,
            'listen_ip': self._listen_ip,
            'listen_port': self._listen_port,
            'req_msg_queue_size': self.req_queue.qsize(),
            # 'res_msg_queue_size': self.res_queue.qsize(),
            'exception_queue_size': self.exception_queue.qsize(),
        }

        return server_info

    def _parse_message(self, conn: socket.socket, message: typing.Any):
        try:
            message = json.loads(message)
        except Exception:
            log.info(f"[SocketAgent][parse_message]message cannot be parsed to json format:{message}")

        if isinstance(message, str) and message.upper() == 'HELLO':
            log.info(f"[SocketAgent][parse_message]hello,world!")

        elif isinstance(message, str) and message.upper() == 'TEST FINISHED':
            self.res_queue.put('TEST FINISHED')

        elif isinstance(message, str) and message.upper() == 'TEST FAILED':
            self.res_queue.put('TEST FAILED')

        elif isinstance(message, str) and message.upper() == 'OPEN DISK':
            control.control_open_value()

        elif isinstance(message, str) and message.upper() == 'RECODE CAPTURE START':
            control.control_recode_capture_start_time()
            if os.path.exists(OSTool.get_current_dir("Test_Report", 'capture_start.txt')):
                log.info('[recode_capture_status]recode capture txt successful')
                return True
            else:
                log.error('[recode_capture_status]recode capture txt fail')
                return False

        elif isinstance(message, str) and message.upper() == 'DEPLOY_PREPARE':
            thin.Thinupdate().test_deploy()

        elif isinstance(message, str) and message.upper() == 'TEST_DEPLOY_2':
            thin.Thinupdate().test_deploy2()


        elif isinstance(message, str) and message.upper() == 'EDITE_U_DISK':
            thin.Thinupdate().test_deploy2()

        elif isinstance(message, str) and message.upper() == 'CHECK CAPTURE STATUS':
            capture_list = []
            deploy_list = []
            deploy_ocr_list = []
            # """3遍capture"""
            for i in range(3):
                time.sleep(60)
                log.info('platfrom value:{}'.format(uut_info_list[1].strip()))
                capture_result = control.control_main_capture_process(platform=uut_info_list[1].strip())
                log.info(f'[check capture status] result value {capture_result}')
                if not capture_result:
                    log.info(f'[check capture status] capture fail and  will restore env')
                    control.control_close_value()
                    control.control_recover_env()
                    time.sleep(600)
                    control.control_switch_user()
                    client = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                    client.send_message('RUN CAPTURE AGAGIN')
                    time.sleep(60)
                    control.control_switch_user()
                    conn1 = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                    conn1.send_message('TEST_CAPTURE_2')
                else:
                    capture_result = control.control_difference_time(start_time_name='capture_start.txt',
                                                                     end_time_name='capture_ended.txt')
                    log.info(
                        '[socket_agent]first {loop} capture result:{capture_result}'.format(loop=len(capture_list) + 1,
                                                                                            capture_result=capture_result))
                    capture_list.append('{}'.format(capture_result))
                    control.control_close_value()
                    time.sleep(5)
                    mac_info = uut_info_list[2].strip().replace(':', '')
                    log.info('mac_info value:{}'.format(mac_info))
                    control.control_wakeup_uut(platform=uut_info_list[1].strip(), mac='{}'.format(mac_info))
                    time.sleep(600)
                    control.control_switch_user()
                    control.control_open_value()
                    log.info('capture_list length:{}'.format(len(capture_list)))
                    if len(capture_list) < 3:
                        conn = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                        conn.send_message('DISABLE WRITE FILTER')
                        time.sleep(60)
                        control.control_switch_user()
                        conn = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                        conn.send_message('TEST_CAPTURE_2')
                        time.sleep(5)
                        control.control_recode_capture_start_time()
                    elif len(capture_list) == 3:
                        conn = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                        conn.send_message('DEPLOY_PREPARE')
                        time.sleep(60)
                        control.control_switch_user()
                        control.control_open_value()
                        conn1 = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                        conn1.send_message('TEST_DEPLOY_2')
                        break
            # """3遍deploy, 如果deploy的图片中出现s，但是OCR VALUE IS None ,请检查ocr中successful识别成什么，然后把字符串复制到control
            # operator 中的deploy_value_list中"""
            for j in range(3):
                mac_info = uut_info_list[2].strip().replace(':', '')
                log.info('mac_info value:{}'.format(mac_info))
                deploy_ocr = control.control_deploy_process(loop=j + 1, platform=uut_info_list[1].strip())
                deploy_result = control.control_difference_time(start_time_name='deploy_start_time.txt',
                                                                end_time_name='deploy_ended_time.txt')
                log.info('[deploy_result] deploy process result value:{a},deploy_ocr value:{b}'.format(b=deploy_result,
                                                                                                       a=deploy_ocr))
                if deploy_result:
                    deploy_ocr_list.append('{}'.format(deploy_ocr))
                    deploy_list.append('{}'.format(deploy_result))
                    control.control_close_value()
                    control.control_wakeup_uut(platform=uut_info_list[1].strip(),
                                               mac='{}'.format(mac_info))
                    if len(deploy_list) < 3:
                        time.sleep(600)
                        control.control_switch_user()
                        conn2 = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                        conn2.send_message('DEPLOY_PREPARE')
                        time.sleep(60)
                        control.control_switch_user()
                        control.control_open_value()
                        time.sleep(5)
                        conn3 = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                        conn3.send_message('TEST_DEPLOY_2')
                        log.info('[deploy_process] this is {} deploy'.format(len(deploy_list)))
                    elif len(deploy_list) == 3:
                        control.control_close_value()
                        time.sleep(600)
                        control.control_switch_user()
                        log.info('control deploy_list value:{}'.format(deploy_list))
                        client = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
                        client.send_message('DISABLE WRITE FILTER')
                        time.sleep(60)
                        control.control_switch_user()
                        break
                else:
                    log.error('deploy_result appear not recover error')
                    return False

            for m in range(3):

                log.info('capture_list value:{}'.format(capture_list))
                log.info('deploy_ocr_value:{}'.format(deploy_ocr_list))
                log.info('deploy_list value:{}'.format(deploy_list))
                vendor = uut_info_list[3].strip()
                print(vendor)
                size = uut_info_list[3].strip().split('-')[1]
                print(size)
                os_info = uut_info_list[4].strip()
                print(os_info)
                ml_info = uut_info_list[5].strip()
                print(ml_info)
                platform_info = uut_info_list[1].strip()
                print(platform_info)
                pn = uut_info_list[-1].strip()
                if pn == 'None':
                    pn = ''
                else:
                    pn = pn
                print(pn)
                save_path = OSTool.get_current_dir('Test_Report',
                                                   'Thinupdate_{a}_{b}.xlsx'.format(a=platform_info, b=os_info))
                thinupdate_result = ['', '', '', '{}'.format(pn), '{}'.format(uut_info_list[1].strip()), '',
                                     '{}'.format(size), '{a}[{b}]'.format(a=os_info, b=ml_info),
                                     '{}'.format(capture_list[m]),
                                     '{}'.format(deploy_ocr_list[m]), '{}'.format(deploy_list[m])]
                log.info('thinupdate_result value:{}'.format(thinupdate_result))

                log.info('thinupdate_result value:{}'.format(thinupdate_result))

                print('template_path value:{}'.format(thin.Thinupdate().template_path))

                if os.path.exists(save_path):
                    data = openpyxl.load_workbook(save_path)
                else:
                    data = openpyxl.load_workbook(thin.Thinupdate().template_path)
                    print(f' data value is {data}')
                sheet = data['Thinupdate']
                max_row = sheet.max_row
                log.info(f'[thin][collect_thin_result] max_row value {max_row}')
                for i in range(len(thinupdate_result)):
                    sheet.cell(max_row + 1, i + 1).value = thinupdate_result[i]
                data.save(save_path)
                data.close()
                time.sleep(3)
            os_info = uut_info_list[4].strip()
            print(os_info)
            platform_info = uut_info_list[1].strip()
            print(platform_info)
            save_path = OSTool.get_current_dir('Test_Report',
                                               'Thinupdate_{a}_{b}.xlsx'.format(a=platform_info, b=os_info))
            control.analyze_report(save_path=save_path)
            receiver = ['sunny.tao@hp.com', 'rengui.li@hp.com']
            additional_email = uut_info_list[-1].strip()
            if additional_email:
                receiver.append(additional_email)
            receiver = list(set(receiver))  # remove duplicate email
            file = control.zip_dir(zipname='{a}_{b}'.format(a=platform_info, b=os_info))
            email_tool.send_mail(recipient=receiver,
                                 subject='Function Performance Test Report',
                                 attachment=file)
            os.remove(file)
            if os.path.exists('uut_info.txt'):
                os.remove('uut_info.txt')
            log.info('[thinupdate] send email successful')
            conn3 = SocketClient(server_ip='127.0.0.1', server_port=55555)
            conn3.send_message('TEST FINISHED')
            return True

        elif isinstance(message, str) and message.upper() == 'DISABLE WRITE FILTER':
            uut.uut_disable_wirte_filter_cmd()
            uut.uut_reboot()

        elif isinstance(message, str) and message.upper() == 'OPEN DISK':
            control.control_open_value()
            log.info('[socket_agent] open u disk to make sure thinupdate can identify usb flash disk')

        elif isinstance(message, str) and message.upper() == 'UUT REBOOT':
            time.sleep(60)
            control.control_open_value()
            control.control_switch_user()
            conn = SocketClient(server_ip='{}'.format(get_uut_ip()), server_port=2333)
            conn.send_message('TEST_CAPTURE_2')
            control.control_recode_capture_start_time()
            conn1 = SocketClient(server_ip='127.0.0.1', server_port=55555)
            conn1.send_message('CHECK CAPTURE STATUS')



        elif isinstance(message, str) and message.upper() == 'RUN CAPTURE AGAGIN':
            thin.Thinupdate().test_capture2_1()

        elif isinstance(message, str) and message.upper() == 'TEST_CAPTURE_2':
            thin.Thinupdate().test_capture2_2()

        elif isinstance(message, dict) and 'thin' in message['name']:
            log.info(f"[SocketAgent][parse_message]start to test {message['name']}")
            thin.start(message)
            log.info(f"[SocketAgent][parse_message]end to test {message['name']}")



        elif isinstance(message, dict) and 'wlan' in message['name']:
            try:
                wlan.start(message)

                if message:
                    try:
                        receiver = ['sunny.tao@hp.com', 'rengui.li@hp.com']
                        if message.get('name', '').lower() == 'wlan':
                            additional_email = message.get('email')
                            additional_tester = message.get('tester')
                            if additional_email:
                                receiver.append(additional_email)
                            if additional_tester:
                                receiver.append(additional_tester)
                        receiver = list(set(receiver))  # remove duplicate email

                        # wlan should be tested alone, and report zip file will be sent to controller t
                        # hat sends email to tester
                        if message['name'] == 'wlan':
                            # send a email to tester after finishing wlan test
                            if int(message['current_config_count']) == int(message['total_config_count']):
                                log.info("[RunTest][run_test]start compress report to zip-format file")
                                file = email_tool.zip_dir()
                                log.info("[RunTest][run_test]zip file path:{}".format(file))

                                # email_tool.send_mail(recipient=receiver,
                                #                      subject='Functional Performance Test Report',
                                #                      attachment=file)
                                # os.remove(file)

                            else:
                                log.info(
                                    f"[RunTest][run_test]current-config/total-config:"
                                    f"{message.get('current_config_count')}/{message.get('total_config_count')}")

                        # else:
                        #     log.info("[RunTest][run_test]start compress report to zip-format file")
                        #     file = email_tool.zip_dir()
                        #     log.info("[RunTest][run_test]zip file path:{}".format(file))
                        #
                        #     email_tool.send_mail(recipient=receiver,
                        #                          subject='Functional Performance Test Report',
                        #                          attachment=file)
                        #     os.remove(file)

                        if OS == 'Windows':
                            for i in range(3):
                                winsound.Beep(300, 1000)
                                time.sleep(0.5)

                    except Exception as e:
                        import traceback
                        log.error(
                            '[SocketAgent][parse_message]meet exception at the end: {}'.format(traceback.format_exc()))

                client = SocketClient(server_ip=get_server_ip(), server_port=6666)
                client.send_message('TEST FINISHED')

            except Exception:
                log.error(
                    f"[SocketAgent][parse_message]receive a exception when running wlan test:{traceback.format_exc()}")

                client = SocketClient(server_ip=get_server_ip(), server_port=6666)
                client.send_message('TEST FAILED')

        elif isinstance(message, dict) and ('boot' in message['name'] or 'reboot' in message['name']):
            log.info(f"[SocketAgent][parse_message]start to test {message['name']}")
            bootstrap.start(message)
            log.info(f"[SocketAgent][parse_message]end to test {message['name']}")

            # client = SocketClient(server_ip=get_server_ip().strip(), server_port=6667)
            # client.send_message('TEST FINISHED')

        else:
            log.error(f"[SocketAgent][run_socket]this message is not supported:{message}")

    def _run_socket(self, backlog: int = 5, keepalive: bool = False, encoding: str = 'utf-8', bufsize: int = 4096):
        try:
            self._socket.bind((self._listen_ip, self._listen_port))
            self._socket.listen(backlog)
            log.info(f"[SocketAgent][run_socket]got socket connection config:\n"
                     f"listen ip:{self._listen_ip}, listen port:{self._listen_port}\n"
                     f"socket connect backlog:{backlog}\n")
        except Exception as e:
            log.error(f"[SocketAgent][run_socket]got exception from socket thread:{traceback.format_exc()}")
            self.exception_queue.put(e)

        while True:
            try:
                if keepalive:
                    log.info("[SocketAgent][run_socket]long connection specification selected")

                    conn, addr = self._socket.accept()
                    log.info(f"[SocketAgent][run_socket]got connection from:{addr}")

                    while True:
                        if not self.req_queue.empty():
                            message = self.req_queue.get()
                            log.info(f"[SocketAgent][run_socket]get message from message queue:{message}")

                            if isinstance(message, str) and message.upper() == 'CLOSE':
                                self.req_queue.task_done()
                                break

                        msg = conn.recv(bufsize).decode(encoding)
                        if len(msg) == 0:
                            continue

                        log.info(f"[SocketAgent][run_socket]got message from {conn.getpeername()}:{msg}")

                        self._parse_message(conn, msg)

                    conn.close()
                    log.info("[SocketAgent][run_socket]long connection closed")

                    break

                else:
                    log.info("[SocketAgent][run_socket]short connection specification selected")

                    conn, addr = self._socket.accept()
                    with conn:
                        log.info(f"[SocketAgent][run_socket]got connection from:{addr}")

                        if not self.req_queue.empty():
                            message = self.req_queue.get()
                            log.info(f"[SocketAgent][run_socket]get message from message queue:{message}")
                            if isinstance(message, str) and message.upper() == 'CLOSE':
                                self.req_queue.task_done()
                                conn.close()
                                log.info("[SocketAgent][run_socket]short connection closed")
                                break

                        msg = conn.recv(bufsize).decode(encoding)
                        if len(msg) == 0:
                            log.warning(f"[SocketAgent][run_socket]got nothing from {conn.getpeername()}")

                        if isinstance(msg, str) and msg.upper() == 'REMOTE_CLOSE':
                            # conn.close()
                            log.info(f"[SocketAgent][run_socket]the server will be closed by remote client: {addr}")
                            break

                        log.info(f"[SocketAgent][run_socket]got message from {conn.getpeername()}:{msg}")

                        self._parse_message(conn, msg)

                    log.info("[SocketAgent][run_socket]short connection closed")

                    # break

            except Exception as e:
                log.error(f"[SocketAgent][run_socket]got exception from socket thread:{traceback.format_exc()}")
                self.exception_queue.put(e)

        log.info(f"[SocketAgent][run_socket]close socket server thread")


class SocketClient:
    def __init__(self, server_ip: str, server_port: int):
        super(SocketClient, self).__init__()

        self._server_ip = server_ip
        self._server_port = server_port

        self._req_msg_queue = _REQ_MSG_QUEUE
        # self._res_msg_queue = _RES_MSG_QUEUE

        self._socket_client: socket.socket = None

    # @property
    # def res_queue(self):
    #     return self._res_msg_queue

    @property
    def req_queue(self):
        return self._req_msg_queue

    def open_client(self, server_ip: str or None = None, server_port: int or None = None) -> socket.socket or None:
        socket_client = None

        if server_ip is not None and server_port is not None:
            self._server_ip = server_ip
            self._server_port = server_port
        try:
            self._socket_client = socket.create_connection((self._server_ip, self._server_port))
            socket_client = self._socket_client

            log.info("[SocketClient][open_client]socket connection created successfully")
        except Exception:
            socket_client = None

            log.error(
                f"[SocketClient][open_client]receive a exception "
                f"when creating socket connection:{traceback.format_exc()}")

        return socket_client

    def close_client(self) -> bool:
        status = False
        try:
            self._socket_client.close()
            status = True

            log.info("[SocketClient][close_client]socket connection closed successfully")
        except Exception:
            status = False

            log.error(
                f"[SocketClient][close_client]receive a exception "
                f"when closing socket connection:{traceback.format_exc()}")

        return status

    def send_message(self, msg: str or dict, keepalive: bool = False, encoding: str = 'utf-8',
                     **kwargs) -> socket.socket or None:
        conn = None

        try:
            if keepalive:
                log.warning(
                    "[SocketClient][send_message]this function cannot support to send message at the long connection")
            else:
                conn = self.open_client(server_ip=kwargs.get('server_ip', self._server_ip),
                                        server_port=kwargs.get('server_port', self._server_port),
                                        # timeout=kwargs.get('timeout', 0.0)
                                        )
                if isinstance(msg, str):
                    msg = bytes(msg, encoding)
                    self._socket_client.sendall(msg)

                    log.info(
                        f"[SocketClient][send_message]send message to server"
                        f"({self._server_ip},{self._server_port}):{msg}")
                elif isinstance(msg, dict):
                    msg = json.dumps(msg, indent=4)
                    msg = bytes(msg, encoding)
                    self._socket_client.sendall(msg)

                    log.info(
                        f"[SocketClient][send_message]send message to server"
                        f"({self._server_ip},{self._server_port}):{msg}")
                else:
                    log.error(f"[SocketClient][send_message]this type is not supported:{type(msg)}")

            return conn

        except Exception:
            log.error(
                f"[SocketClient][send_message]receive a exception "
                f"when sending message({msg}):{traceback.format_exc()}")

    # def receive_message(self, kwargs) -> typing.Any:
    #     msg = None
    #     try:
    #         if not self.res_queue.empty():
    #             msg = self.res_queue.get(block=kwargs.get('block', False),
    #                                      timeout=kwargs.get('timeout', 0.0))
    #         else:
    #             log.warning("[SocketClient][receive_message]response message queue is empty currently")
    #     except Exception:
    #         log.error(
    #             f"[SocketClient][receive_message]receive a exception "
    #             f"when receiving message:{traceback.format_exc()}")
    #
    #     return msg

    def get_client_info(self) -> dict:
        client_info = {
            'server_ip': self._server_ip,
            'server_port': self._server_port,
            # 'req_msg_queue_size': self.req_queue.qsize(),
            # 'res_msg_queue_size': self.res_queue.qsize(),
        }

        return client_info
