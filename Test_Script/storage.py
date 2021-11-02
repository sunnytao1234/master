# -*- coding: utf-8 -*-
# @time     :   4/1/2021 3:36 PM
# @author   :   balance
# @file     :   storage.py
import os
import shutil
import subprocess
import time

import openpyxl
from openpyxl.styles import PatternFill

from Common import common
from Common.common import get_diskmark_version, alter
from Common.common import get_element
from Common.common_function import get_current_dir
from Common.exception import IncorrectOS
from Common.log import log


class StorageWin:
    exe_file_32 = r'C:\Program Files\{}\DiskMark32.exe'.format(get_diskmark_version().strip().split('_')[0])
    exe_file_64 = r'C:\Program Files\{}\DiskMark64.exe'.format(get_diskmark_version().strip().split('_')[0])

    def __init__(self, status):
        """storage test for wes"""
        self.__test_data = None
        self.status = status
        self.save_path = get_current_dir('result.xlsx')
        self.template_path = get_current_dir('Test_Data', 'excle_template.xlsx')
        self.data_start_col = 10
        self.data_start_row = 3
        """
        {name: 'storage', loop: 3, os: wes, platform: mt31, tester: balance@hp.com, 
        os_info: windows10-xx-xxx, platform_info: hp mt31, ml_info: xxxxx, memory_info: xxxx8G,
        disk_info: sumsun xxxxx32G, disk_pn: 12345, bios_info: xxxx, cpu_info: xxxx@2.0GHz,
        gpu_info: xxxx, main_board_info: 321a}
        """

    def install(self):
        """ install passmark in uut
        1. file store in Test_utility
        2. install with quiet mode
        """

        softpath = get_current_dir('Test_Utility',
                                   '{diskmark_version}.exe'.format(diskmark_version=get_diskmark_version()))
        subprocess.Popen('{} /silent'.format(softpath))
        time.sleep(25)
        if os.path.exists(self.exe_file_64):
            log.info("[storage][install]install software success")
            return True
        else:
            log.info("install software fail")
            subprocess.Popen('{} /silent'.format(softpath))

    def launch(self, wait_time: int = 20):
        """
        launch diskmark
        :return:
        """
        copy_target_path = r'C:\Program Files\{}\DiskMark64.ini'.format(get_diskmark_version().strip().split('_')[0])
        diskmark_path = get_current_dir('Test_Utility', 'DiskMark64.ini')
        shutil.copy(diskmark_path, copy_target_path)
        log.info('[storage][launch]begin to launch diskmark ')
        subprocess.Popen(self.exe_file_64)
        time.sleep(5)
        if common.wait_element(common.get_element('STORAGE_WND'), 5):
            get_element('STORAGE_ALL').Click()
            # wait to launch
            time.sleep(wait_time)

            return True
        else:
            return False

    def set_config(self, value: dict, **kwargs):
        self.__test_data = value
        self.save_path = get_current_dir('Test_Report',
                                         'Storage_{}_{}.xlsx'.format(value.get('platform_info'),
                                                                     value.get('os_info')))

    def check_status(self):
        """
        get diskmark test status
        :return: status of diskmark
        """
        status = ''
        value = common.get_element(
            'STORAGE_MENU').GetParentControl().GetParentControl().Name
        log.info("[storage][check status]this value is {}".format(value))
        if 'CrystalDiskMark 8.0.2 x64 [Admin]' in value:
            self.status['status'] = 'Complete'
            status = 'Complete'
        else:
            self.status['status'] = status
            status = value
        return status

    def stop(self, judge='True'):
        """
        :param judge: True or False to choose stop method
        :return:
        """
        if judge == True:
            get_element('STORAGE_STOP').Click()
        else:
            subprocess.Popen('taskkill /F /IM DiskMark8.exe')
        return True

    # @common.timer('collect result')
    def __collect_result(self):
        """
        get test result and record to specific excel file
        :return: list
        return: [seq32_read, 4k32_read, seq_read, 4k_read, seq32_write, 4k32_write, seq_write, 4k_write]
        """

        strage_value = []
        table = get_element('STORAGE_MENU')
        length = len(table.GetParentControl().GetParentControl().GetChildren())
        for i in range(length):
            write_value = table.GetParentControl().GetParentControl().GetChildren()[i].Name.strip()
            strage_value.append(write_value)
            for j in strage_value:
                if j == '':
                    strage_value.remove('')
        test_value = strage_value[5:13]
        return test_value

    # @common.timer('export result to excel')
    def __export_result_to_excle(self):
        """export collect data to excle"""
        try:
            config_list = ['', self.__test_data.get('disk_info'),
                           self.__test_data.get('disk_info').split()[-1], self.__test_data.get('disk_pn'),
                           self.__test_data.get('platform_info'),
                           '', '', '',
                           '{}[{}]'.format(self.__test_data.get('os_info'), self.__test_data.get('ml_info'))]
        except Exception:
            config_list = ['', self.__test_data.get('disk_info'),
                           '', self.__test_data.get('disk_pn'),
                           self.__test_data.get('platform_info'),
                           '', '', '',
                           '{}[{}]'.format(self.__test_data.get('os_info'), self.__test_data.get('ml_info'))]

        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)
        sheet = data['Storage']
        max_row = sheet.max_row
        log.debug('[storage][export excel]current used rows {}'.format(max_row))

        test_result = self.__collect_result()
        config_list.extend(test_result)
        for i in range(len(config_list)):
            sheet.cell(max_row + 1, i + 1).value = config_list[i]
        data.save(self.save_path)
        data.close()
        time.sleep(3)
        return True

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []
        for item in result:
            if abs(item.get('value') - average) / item.get('value') > deviation:
                abnormal_positions.append(item.get('position'))
                log.debug('[storage][get abnormal data]Find abnormal data cell: {}'.format(item.get('position')))
        return abnormal_positions

    def analyze_report(self):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 10, row 1
        :return:
        """
        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)
        sheet = data['Storage']
        max_row = sheet.max_row
        log.debug(
            '[storage wes][analyze_report]current used rows {}, start row {}'.format(max_row, self.data_start_row))
        for col in range(self.data_start_col, self.data_start_col + 8):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
            for row in range(self.data_start_row, max_row + 1):
                cell_value = sheet.cell(row, col).value
                cell_value = float(cell_value) if cell_value is not None else 0
                sum_data += cell_value
                temp_data = {'position': (row, col), 'value': cell_value}
                result.append(temp_data)

            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row + 1 - self.data_start_row)

            sheet.cell(max_row + 1, col).value = average_data
            abnormals = self.__get_abnormal_data(result, average_data)
            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill
        data.save(self.save_path)
        data.close()

    def test_storage(self):
        """
        run using specific config
        :return:
        """
        for loop in range(self.__test_data.get('loop')):
            # for loop in range(3):
            try:
                self.launch()
            except:
                log.error('[storage wes][test storage]fail to launch diskmark')
            common.get_element('STORAGE_ALL').Click()
            for t in range(500):
                # test will cost 5min, give 120*5 = 10min to check status until finished
                status = self.check_status()
                if status == 'Complete':
                    # check not exist will cost 13s more or less
                    log.info(f'[storage wes][test storage]Test cycle {loop + 1} Finished')
                    self.__export_result_to_excle()
                    self.status['status'] = f'{status}_loop {int(loop) + 1}'
                    break
                else:
                    self.status['status'] = f'{status}_loop {int(loop) + 1}'
                    log.debug(f'[storage wes][test storage]testing {status} of loop {int(loop) + 1}')
                    time.sleep(10)
            get_element("STORAGE_STOP").Click()


class StorageLinux:
    def __init__(self, status):
        """storage test for linux"""
        self.__test_data = None
        self.status = status

        self.save_path = get_current_dir('result.xlsx')
        self.template_path = get_current_dir('Test_Data', 'excle_template.xlsx')

        self.data_start_col = 10
        self.data_start_row = 3

    def install(self):
        alter(get_current_dir('/etc/apt/sources.list'), '#', '')
        time.sleep(3)
        os.system('sudo apt-get -o Acquire::http::proxy=http://15.85.199.199:8080 update')
        os.popen('fsunlock')
        os.system('apt-get -o Acquire::http::proxy=http://15.85.199.199:8080 install -y fio')
        time.sleep(20)
        check_out = subprocess.getoutput('dpkg -l | grep fio')
        log.info("check out value :{}".format(check_out))
        if check_out == '':
            log.error("[linux storage][install] install fail")
            return False
        else:
            log.info("[linux storage][install] install successful")
            return True

    def launch(self):
        """
        launch storage test command
        :return:
        """
        log.info('begin to test storage linux')
        return True

    def set_config(self, value: dict, **kwargs):
        self.__test_data = value
        self.save_path = get_current_dir('Test_Report',
                                         'Storage_{}_{}.xlsx'.format(value.get('platform_info'),
                                                                     value.get('os_info')))

    def check_status_and_collect_result(self):
        status = ''
        cmd = 'fio -filename=/test -direct=1 -iodepth 32 -thread -rw=read -ioengine=libaio -bs=1m -size=1G ' \
              '-numjobs=1 -runtime=30 -group_reporting -name=read_1M_Q32T1 && rm /test &&fio -filename=/test -direct=1 -iodepth 32 -thread -rw=randread -ioengine=libaio -bs=4k -size=1G ' \
              '-numjobs=8 -runtime=30 -group_reporting -name=randread_4k_Q32T8&& rm /test &&fio -filename=/test -direct=1 -iodepth 1 -thread -rw=read -ioengine=libaio -bs=1m -size=1G ' \
              '-numjobs=1 -runtime=30 -group_reporting -name=read_1M_Q1T1 && rm /test &&fio -filename=/test -direct=1 -iodepth 1 -thread -rw=randread -ioengine=libaio -bs=4k -size=1G ' \
              '-numjobs=1 -runtime=30 -group_reporting -name=randread_4k_Q32T1 && rm /test &&fio -filename=/test -direct=1 -iodepth 32 -thread -rw=write -ioengine=libaio -bs=1m -size=1G ' \
              '-numjobs=1 -runtime=30 -group_reporting -name=write_1M_Q32T1 && rm /test &&fio -filename=/test -direct=1 -iodepth 32 -thread -rw=randwrite -ioengine=libaio -bs=4k -size=1G ' \
              '-numjobs=8 -runtime=30 -group_reporting -name=randwrite_4k_Q32T8 && rm /test &&fio -filename=/test -direct=1 -iodepth 1 -thread -rw=write -ioengine=libaio -bs=1m -size=1G ' \
              '-numjobs=1 -runtime=30 -group_reporting -name=write_1M_Q1T1 && rm /test &&fio -filename=/test -direct=1 -iodepth 1 -thread -rw=randwrite -ioengine=libaio -bs=4k -size=1G ' \
              '-numjobs=1 -runtime=30 -group_reporting -name=randwrite_4k_Q32T1 && rm /test'
        result_list = []
        result = subprocess.getoutput(cmd)
        log.info("result value is:{}".format(result))
        if 'randwrite_4k_Q32T1' in result:
            self.status['status'] = 'Complete'
            status = 'Complete'
        value_list = result.split('Run status group 0 (all jobs):')
        log.info("value is:{a},type is:{b}".format(a=value_list, b=type(value_list)))
        length = len(value_list)
        for i in range(1, length):
            read_write_value = value_list[i].strip().split('(')[1].split('MB/s)')[0]
            result_list.append(read_write_value)
        log.info("[linux-check status and collect result] value is:{}".format(result_list))
        return result_list, status

    def export_result_to_excle(self):
        """export collect data to excle"""
        try:
            config_list = ['', self.__test_data.get('disk_info'),
                           self.__test_data.get('disk_info').split()[-1], self.__test_data.get('disk_pn'),
                           self.__test_data.get('platform_info'),
                           '', '', '',
                           '{a}[{b}]'.format(a=self.__test_data.get('os_info'),
                                             b=subprocess.getoutput('cat /etc/imageid'))
                           ]
        except Exception:
            config_list = ['', self.__test_data.get('disk_info'),
                           '', self.__test_data.get('disk_pn'),
                           self.__test_data.get('platform_info'),
                           '', '', '',
                           '{a}[{b}]'.format(a=self.__test_data.get('os_info'),
                                             b=subprocess.getoutput('cat /etc/imageid'))
                           ]

        test_result, test_status = self.check_status_and_collect_result()
        config_list.extend(test_result)
        log.info("config list value is:{}".format(config_list))

        if os.path.exists(self.save_path):
            data = openpyxl.load_workbook(self.save_path)
        else:
            data = openpyxl.load_workbook(self.template_path)

        sheet = data['Storage']
        max_row = sheet.max_row
        log.info("max row value is :{}".format(max_row))
        for i in range(len(config_list)):
            sheet.cell(max_row + 1, i + 1).value = config_list[i]
        data.save(self.save_path)
        data.close()
        time.sleep(3)
        return test_status

    def __get_abnormal_data(self, result: list, average: float, deviation=0.05) -> list:
        """
        get average value
        :param result: [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
        :param average: average of value in result
        :param deviation: Deviation of single data
        :return: position list of Deviation :[(row1, col1), (row2, col2)..]
        """
        abnormal_positions = []
        for item in result:
            if abs(item.get('value') - average) / item.get('value') > deviation:
                abnormal_positions.append(item.get('position'))
                log.debug('[storage][get abnormal data]Find abnormal data cell: {}'.format(item.get('position')))
        return abnormal_positions

    def test_storage(self):
        """
        run using specific config
        :return:
        """
        for loop in range(self.__test_data.get('loop')):
            status = self.export_result_to_excle()
            if status == 'Complete':
                log.info(f'[storage linux][test storage]Test cycle {loop + 1} Finished')
                self.status['status'] = f'{status}_loop {int(loop) + 1}'
            else:
                log.error(f'[storage linux][test storage]Test cycle {loop + 1} Finished')
                self.status['status'] = f'{status}_loop {int(loop) + 1}'
                return False

    def analyze_report(self):
        """
        1. get average data in excel
        2. mark abnormal data cell as red
        test data start from col 12, row 1
        :return:
        """
        red_fill = openpyxl.styles.PatternFill("solid", fgColor="FF0000")
        data = openpyxl.load_workbook(self.save_path)
        sheet = data['Storage']
        max_row = sheet.max_row
        log.debug(
            '[storage wes][analyze_report]current used rows {}, start row {}'.format(max_row, self.data_start_row))
        for col in range(self.data_start_col, self.data_start_col + 8):
            sum_data = 0
            result = []  # [{'position': (row, col), 'value': xx},{'position': (row2, col2), 'value': xx2}]
            for row in range(self.data_start_row, max_row + 1):
                cell_value = sheet.cell(row, col).value
                cell_value = float(cell_value) if cell_value is not None else 0
                sum_data += cell_value
                temp_data = {'position': (row, col), 'value': cell_value}
                result.append(temp_data)
            if sum_data == 0:
                average_data = 0
            else:
                average_data = sum_data / (max_row + 1 - self.data_start_row)
            sheet.cell(max_row + 1, col).value = average_data
            abnormals = self.__get_abnormal_data(result, average_data)
            for abnormal in abnormals:
                sheet.cell(abnormal[0], abnormal[1]).fill = red_fill
        data.save(self.save_path)
        data.close()


def start(value: dict, **kwargs):
    """
    main function for storage test
    1. support cross platform
    2. support all storage test items
    3. collect result and analyze test result
       3.1 get result
       3.2 get average result
       3.3 mark abnormal data as red in excel
    4. write to excel as expected format
    5. all exception should be covered
    :param value:
    :param kwargs:
    :return: None
    """
    status = kwargs.get('status')
    try:
        if value.get('os') == 'wes':
            storage = StorageWin(status)
        elif value.get('os') == 'linux':
            storage = StorageLinux(status)
        else:
            log.error('[storage][start]Incorrect OS given')
            raise IncorrectOS('Incorrect OS given')
        storage.install()
        storage.set_config(value)
        storage.test_storage()
        storage.analyze_report()
        time.sleep(10)
    except:
        import traceback
        traceback.print_exc()
        log.error('[storage][start]Get exception during test storage')
