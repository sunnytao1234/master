# -*- coding: utf-8 -*-
# @time     :   4/21/2021 4:39 PM
# @author   :   balance
# @file     :   reboot.py
# todo:develop and debug
import os
import statistics
import time
import typing

import openpyxl
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet

from Common import email_tool
from Common.common_function import get_current_dir
from Common.log import log
from Test_Script.base import socket_action
from Test_Script.base.bootstrap import BaseBootTest, ReBootResult

log.log_level = 4


class WesRebootTest(BaseBootTest):
    def __init__(self, config: dict):
        super(WesRebootTest, self).__init__(config=config,
                                            platform=config['platform'],
                                            boot_wait_time=config['boot_wait_time'],
                                            camera_device_index=config['camera_device_index'],
                                            manual_boot_list=config['manual_boot_list'])

    def test_result(self) -> typing.List[ReBootResult]:
        log.debug("[WesRebootTest][test_result]the test_result process is started")

        results = []

        for cycle in range(self.config['loop']):
            log.info(f"[WesRebootTest][test_result]current test progress: [{cycle + 1}/{self.config['loop']}]")

            if self.observer.check_desktop():
                self.timer.set_time_report(desktop_time=time.time())
                if self.observer.reboot():
                    if self.observer.check_shutdown():
                        self.timer.set_time_report(shutdown_time=time.time())
                        if self.observer.check_hp_logo():
                            self.timer.set_time_report(hp_logo_time=time.time())
                            if self.observer.check_desktop():
                                self.timer.set_time_report(next_desktop_time=time.time())
                                time.sleep(45)  # wait to start socket agent
                            else:
                                log.error(f"[WesRebootTest][test_result]desktop does not appear")
                        else:
                            log.error(f"[WesRebootTest][test_result]hp logo does not appear")
                    else:
                        log.error(f"[WesRebootTest][test_result]cannot shutdown")
                else:
                    log.error(f"[WesRebootTest][test_result]cannot reboot this uut")
            else:
                log.error(f"[WesRebootTest][test_result]desktop does not appear")

            results.append(self.timer.get_time_report())

            log.info(f"[WesRebootTest][test_result]please wait to {self.boot_wait_time}s")
            time.sleep(self.boot_wait_time)

        log.debug("[WesRebootTest][test_result]the test_result process is ended")

        return results

    def analyse_result(self, results: typing.List[ReBootResult]) -> typing.Tuple[
        typing.List, typing.List, typing.List, typing.List, typing.List]:
        log.debug("[WesRebootTest][analyse_result]the analyse_result process is started")

        os_shutdown_times, bios_hot_post_times, startup_times, total_times, average_times = [], [], [], [], []

        for index, row_data in enumerate(results):
            os_shutdown_times.append(row_data.shutdown_time - row_data.desktop_time)
            bios_hot_post_times.append(row_data.hp_logo_time - row_data.shutdown_time)
            startup_times.append(row_data.next_desktop_time - row_data.hp_logo_time)
            total_times.append(row_data.next_desktop_time - row_data.desktop_time)

        average_times.append(statistics.mean(os_shutdown_times))
        average_times.append(statistics.mean(bios_hot_post_times))
        average_times.append(statistics.mean(startup_times))
        average_times.append(statistics.mean(total_times))

        log.debug("[WesRebootTest][analyse_result]the analyse_result process is ended")

        return os_shutdown_times, bios_hot_post_times, startup_times, total_times, average_times

    def write_result(self, results: typing.Tuple[typing.List, typing.List, typing.List, typing.List, typing.List],
                     template_path: str = get_current_dir('Test_Data', 'excle_template.xlsx'),
                     fill_color=styles.PatternFill("solid", fgColor="FF0000"),
                     abnormal_threshold: float = 0.05
                     ):
        log.debug("[WesRebootTest][write_result]the write_result process is started")

        save_path = get_current_dir('Test_Report',
                                    'Boot_{}_{}.xlsx'.format(self.config['platform'], self.config['os']))

        if os.path.exists(save_path):
            excel = openpyxl.load_workbook(save_path)
        else:
            excel = openpyxl.load_workbook(template_path)

        try:
            base_info = [
                '', self.config['disk_info'], self.config['disk_info'].split()[-1], '',
                self.config['platform_info'], self.config['memory_info'], self.config['cpu_info'],
                self.config['gpu_info'],
                '', self.config['bios_info'], self.config['os_info'] + self.config['ml_info'], self.config['ml_info'],
            ]
        except Exception:
            base_info = [
                '', self.config['disk_info'], '', '',
                self.config['platform_info'], self.config['memory_info'], self.config['cpu_info'],
                self.config['gpu_info'],
                '', self.config['bios_info'], self.config['os_info'] + self.config['ml_info'], self.config['ml_info'],
            ]

        boot_sheet: Worksheet = excel['Booting Time ']
        current_max_row, current_max_col = boot_sheet.max_row, boot_sheet.max_column

        boot_sysinfo_cells = boot_sheet['A{}'.format(current_max_row - 3):'L{}'.format(current_max_row)]
        log.info(
            f"[WesRebootTest][write_result]current boot system info cells: "
            f"({'A{}'.format(current_max_row - 3)}:{'L{}'.format(current_max_row)})")

        boot_result_cells = boot_sheet['P{}'.format(current_max_row - 3):'S{}'.format(current_max_row)]
        log.info(
            f"[WesRebootTest][write_result]current boot time result cells: "
            f"({'P{}'.format(current_max_row - 3)}:{'S{}'.format(current_max_row)})")

        for boot_sysinfo_cell_row in boot_sysinfo_cells:
            for index, cell in enumerate(boot_sysinfo_cell_row):
                cell.value = base_info[index]

        for row, boot_result_cell_row in enumerate(boot_result_cells):
            if row == len(boot_result_cells) - 1:
                for col, cell in enumerate(boot_result_cell_row):
                    cell.value = round(results[-1][col], 2)
            else:
                for col, cell in enumerate(boot_result_cell_row):
                    cell.value = round(results[col][row], 2)
                    average_value = round(results[-1][col], 2)
                    log.info(f"[WesRebootTest][write_result]current average value is: {average_value}")

                    if average_value == 0:
                        cell.fill = fill_color
                        log.info(f"[WesRebootTest][write_result]abnormal value is funded:{cell}")

                    elif abs(cell.value - average_value) / average_value > abnormal_threshold:
                        cell.fill = fill_color
                        log.info(f"[WesRebootTest][write_result]abnormal value is funded:{cell}")

        excel.save(save_path)
        log.info(f"[WesRebootTest][write_result]result excel save to:{save_path}")

        log.debug("[WesRebootTest][write_result]the write_result process is ended")


def start(value: dict):
    log.debug("[reboot][start]the reboot test process is started")

    if value['os'] not in ('wes', 'linux'):
        log.error(f'[reboot][start]a incorrect os argument is given:{value["os"]}')
    else:
        with WesRebootTest(config=value) as reboot:
            result = reboot.test_result()
            result = reboot.analyse_result(result)
            reboot.write_result(result)

        log.info(f"[reboot][start]current reboot-time result: {result}")

        receiver = ['sunny.tao@hp.com', 'rengui.li@hp.com', value['tester']]
        receiver = list(set(receiver))

        log.info(f"[reboot][start]receivers are:{receiver}")

        file = email_tool.zip_dir()
        log.info("[reboot][start]zip file path:{}".format(file))

        email_tool.send_mail(recipient=receiver,
                             subject='Functional Performance Test Report',
                             attachment=file)
        os.remove(file)
        log.info("[reboot][start]send emails successfully")

    log.debug("[reboot][start]the reboot test process is ended")
